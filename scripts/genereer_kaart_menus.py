#!/usr/bin/env python3
"""
Genereer de kaart-menu-pagina's uit vizier.xlsx.

Voor elke taal (nl/de/en):
  Lees de 19 kaart-verwijzingen uit code x.1.3.1.3.5 t/m x.1.3.1.3.23
  Bouw een menu-pagina met tegels naar elke kaart
  Schrijf naar:
    /nl/stemgedrag.html
    /de/wahlfolgen.html
    /en/vote-impact.html

De matrix (nl/_data/vizier.xlsx) is de canonieke bron. Dit script verzint
geen data — alle namen, URLs en volgorde komen uit de matrix.

Gebruik:
  python3 scripts/genereer_kaart_menus.py

Wordt uitgevoerd door GitHub Action (xlsx-naar-json.yml) bij elke push
die vizier.xlsx wijzigt.
"""
import openpyxl
from pathlib import Path
import re
import sys

REPO = Path(__file__).resolve().parent.parent
XLSX = REPO / "nl" / "_data" / "vizier.xlsx"

# Configuratie per taal
TAAL_CFG = {
    'nl': {
        'lang_prefix': '1',
        'menu_bestand': REPO / 'nl' / 'stemgedrag.html',
        'html_lang': 'nl',
        'titel_kaart': 'Gevolgenkaart',
        'titel_page': 'Gevolgenkaart — Het Open Vizier',
        'ondertitel': 'Partij × Persoon × Basislijn = Projectie',
        'h1_intro': 'Een matrixen-bestand als spiegel van de politiek',
        'kies_kop': 'Kies een landsversie',
        'kies_intro': 'Elke landsversie draait op dezelfde 84 regels; alleen partijlijsten en basislijn verschillen. Klik op een land om de app te openen.',
        'app_label': 'Gevolgenkaart',
        'voorpagina_link': 'Voorpagina',
    },
    'de': {
        'lang_prefix': '2',
        'menu_bestand': REPO / 'de' / 'wahlfolgen.html',
        'html_lang': 'de',
        'titel_kaart': 'Folgenkarte',
        'titel_page': 'Folgenkarte — Das Offene Visier',
        'ondertitel': 'Partei × Person × Basislinie = Projektion',
        'h1_intro': 'Eine Matrix-Datei als Spiegel der Politik',
        'kies_kop': 'Wählen Sie eine Landesversion',
        'kies_intro': 'Jede Landesversion läuft auf denselben 84 Regeln; nur Parteilisten und Basislinie unterscheiden sich. Klicken Sie auf ein Land, um die App zu öffnen.',
        'app_label': 'Folgenkarte',
        'voorpagina_link': 'Startseite',
    },
    'en': {
        'lang_prefix': '3',
        'menu_bestand': REPO / 'en' / 'vote-impact.html',
        'html_lang': 'en',
        'titel_kaart': 'Consequence Map',
        'titel_page': 'Consequence Map — The Open Visor',
        'ondertitel': 'Party × Person × Baseline = Projection',
        'h1_intro': 'A matrix file as mirror of politics',
        'kies_kop': 'Choose a country version',
        'kies_intro': 'Each country version runs on the same 84 rules; only party lists and baseline differ. Click a country to open the app.',
        'app_label': 'Consequence Map',
        'voorpagina_link': 'Home',
    },
}

# Vlag-emoji per land-code
VLAG_PER_LAND = {
    'nl': '🇳🇱', 'de': '🇩🇪', 'ch': '🇨🇭', 'es': '🇪🇸',
    'eu': '🇪🇺', 'us': '🇺🇸', 'cn': '🇨🇳', 'in': '🇮🇳',
    'kr': '🇰🇷', 'jp': '🇯🇵', 'tw': '🇹🇼', 'se': '🇸🇪',
    'il': '🇮🇱', 'ir': '🇮🇷', 'za': '🇿🇦', 'br': '🇧🇷',
    'cu': '🇨🇺', 'sr': '🇸🇷',
}

# Hero-afbeelding per land (bestaande naming-convention in ../assets/wat-opkomt/)
HERO_PER_LAND = {
    'nl': 'H_nederland_industrie_zakt.jpg',
    'de': 'H_de_wahlfolgen.jpg',
    'ch': 'H_ch_folgenkarte.jpg',
    'es': 'H_es_mapa.jpg',
    'eu': 'H_brussel_gevolgenkaart.jpg',
    'us': 'H_us_consequence.jpg',
    'cn': 'H_cn_gevolgenkaart.jpg',
    'in': 'H_in_gevolgenkaart.jpg',
    'kr': 'H_kr_gevolgenkaart.jpg',
    'jp': 'H_jp_gevolgenkaart.jpg',
    'tw': 'H_tw_gevolgenkaart.jpg',
    'se': 'H_se_gevolgenkaart.jpg',
    'il': 'H_il_gevolgenkaart.jpg',
    'ir': 'H_ir_gevolgenkaart.jpg',
    'za': 'H_za_gevolgenkaart.jpg',
    'br': 'H_br_gevolgenkaart.jpg',
    'cu': 'H_cu_gevolgenkaart.jpg',
    'sr': 'H_sr_gevolgenkaart.jpg',
}


def extract_land_from_url(url):
    """Uit '../nl/gevolgenkaart-cu/' haal 'cu'."""
    m = re.search(r'-([a-z]{2})/?$', url.rstrip('/'))
    return m.group(1) if m else None


def lees_kaarten_uit_matrix(ws, lang_prefix):
    """Lees alle kaart-verwijzingen voor een taal.

    De 19 kaarten zitten op codes {lang_prefix}.1.3.1.3.5 t/m .23.
    """
    kaarten = []
    for r in range(3, ws.max_row + 1):
        code = ws.cell(r, 2).value
        if not code:
            continue
        code_str = str(code)
        parts = code_str.split('.')
        if len(parts) != 6:
            continue
        if parts[0] != lang_prefix:
            continue
        if parts[1:5] != ['1', '3', '1', '3']:
            continue
        try:
            n = int(parts[5])
        except ValueError:
            continue
        if not (5 <= n <= 23):
            continue

        naam = ws.cell(r, 6).value or ''
        url = ws.cell(r, 9).value or ''
        ondertitel = ws.cell(r, 7).value or ''
        if not url.strip():
            continue

        land = extract_land_from_url(url)
        if not land:
            continue

        kaarten.append({
            'code': code_str,
            'volgorde': n,
            'naam': str(naam).strip(),
            'ondertitel': str(ondertitel).strip(),
            'url': str(url).strip(),
            'land': land,
        })
    kaarten.sort(key=lambda k: k['volgorde'])
    return kaarten


def genereer_menu_html(kaarten, cfg):
    tegels = []
    for k in kaarten:
        vlag = VLAG_PER_LAND.get(k['land'], '')
        hero = HERO_PER_LAND.get(k['land'], f"H_{k['land']}_gevolgenkaart.jpg")
        ondertitel = k['ondertitel'] or ''
        tegel = f'''    <a class="app-button" href="{k['url']}">
      <img class="hero" src="../assets/wat-opkomt/{hero}" alt="" loading="lazy">
      <div class="tegel-body">
        <span class="flag">{vlag}</span>
        <span class="naam">{k['naam']}</span>
        <span class="app-label">{cfg['app_label']}</span>
        <span class="toelichting">{ondertitel}</span>
      </div>
    </a>'''
        tegels.append(tegel)
    tegels_html = '\n'.join(tegels)

    html = f'''<!DOCTYPE html>
<html lang="{cfg['html_lang']}">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{cfg['titel_page']}</title>
<meta name="description" content="{cfg['ondertitel']}">
<link rel="stylesheet" href="../assets/style.css">
<link rel="icon" type="image/svg+xml" href="../assets/favicon.svg">
<meta name="generator" content="scripts/genereer_kaart_menus.py — bron: nl/_data/vizier.xlsx">
<style>
  .apps-sectie {{ max-width:1100px; margin:2.5rem auto 3rem; padding:0 1.5rem; }}
  .apps-sectie h2 {{ font-family:Georgia,serif; font-size:1.5rem; color:#1a1a1a; text-align:center; margin:0 0 .5rem; }}
  .apps-sectie .intro {{ text-align:center; font-family:Georgia,serif; font-style:italic; color:#4a5263; max-width:640px; margin:0 auto 1.6rem; line-height:1.55; }}
  .apps-grid {{ display:grid; grid-template-columns:repeat(auto-fill,minmax(300px,1fr)); gap:1rem; max-width:960px; margin:0 auto; }}
  .app-button {{ display:flex; align-items:stretch; background:#1c5760; color:#f5f0e6; text-decoration:none; border-radius:6px; overflow:hidden; transition:background .2s ease, transform .15s ease; box-shadow:0 4px 14px rgba(0,0,0,.08); min-height:150px; }}
  .app-button:hover {{ background:#164a52; transform:translateY(-2px); }}
  .app-button .hero {{ width:150px; height:150px; object-fit:cover; flex-shrink:0; border-right:1px solid rgba(255,255,255,.15); }}
  .app-button .tegel-body {{ padding:.9rem 1rem; display:flex; flex-direction:column; justify-content:center; min-width:0; }}
  .app-button .flag {{ font-size:1.4rem; margin-bottom:.2rem; line-height:1; }}
  .app-button .naam {{ font-family:Georgia,serif; font-weight:700; font-size:1.1rem; margin-bottom:.15rem; line-height:1.2; }}
  .app-button .app-label {{ font-size:.78rem; letter-spacing:.06em; text-transform:uppercase; opacity:.75; margin-bottom:.35rem; font-weight:600; }}
  .app-button .toelichting {{ font-size:.82rem; opacity:.88; line-height:1.35; }}
  @media (max-width:400px) {{
    .app-button {{ flex-direction:column; }}
    .app-button .hero {{ width:100%; height:120px; border-right:none; border-bottom:1px solid rgba(255,255,255,.15); }}
  }}
</style>
</head>
<body>

<header class="masthead" style="text-align:center; padding:2rem 1.5rem 0.8rem;">
  <p class="masthead__kicker" style="font-size:0.78rem; letter-spacing:0.15em; text-transform:uppercase; color:#1c5760; margin:0 0 0.6rem 0; font-weight:600;">Het Open Vizier</p>
  <a class="masthead__logo" href="./" style="display:inline-block; font-family:Georgia,serif; font-style:italic; font-weight:700; font-size:clamp(2rem,4.5vw,3rem); color:#1a1a1a; text-decoration:none;">{cfg['titel_kaart']}</a>
  <p class="masthead__sub" style="font-family:Georgia,serif; font-style:italic; color:#4a5263; margin:0.4rem 0 0; font-size:1.05rem;">{cfg['ondertitel']}</p>
</header>

<nav class="nav">
  <div class="nav__inner">
    <ul class="nav__links">
      <li><a href="./">{cfg['voorpagina_link']}</a></li>
    </ul>
  </div>
</nav>

<article style="padding:2rem 0 1rem;">
  <div style="max-width:780px; margin:0 auto; padding:0 1.5rem;">
    <h1 style="font-family:Georgia,serif; font-size:clamp(1.7rem,3.2vw,2.3rem); line-height:1.15; margin:0 0 1rem; color:#1a1a1a; text-align:center;">{cfg['h1_intro']}</h1>
  </div>
</article>

<section class="apps-sectie">
  <h2>{cfg['kies_kop']}</h2>
  <p class="intro">{cfg['kies_intro']}</p>
  <div class="apps-grid">
{tegels_html}
  </div>
</section>

</body>
</html>
'''
    return html


def main():
    if not XLSX.exists():
        print(f"FOUT: {XLSX} bestaat niet", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb['Knopen']

    for taal, cfg in TAAL_CFG.items():
        kaarten = lees_kaarten_uit_matrix(ws, cfg['lang_prefix'])
        if len(kaarten) != 19:
            print(f"WAARSCHUWING: taal '{taal}' heeft {len(kaarten)} kaarten in matrix (verwacht: 19)")
        html = genereer_menu_html(kaarten, cfg)
        cfg['menu_bestand'].write_text(html, encoding='utf-8')
        print(f"✓ {cfg['menu_bestand'].relative_to(REPO)}: {len(kaarten)} kaarten geschreven uit matrix")


if __name__ == '__main__':
    main()
