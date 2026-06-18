#!/usr/bin/env python3
"""Converteer vizier.xlsx naar de 5 JSON-tabellen die de website leest.

Workflow:
1. Lees nl/_data/vizier.xlsx (tabblad 'Knopen' + 'Layouts')
2. Schrijf 5 tabellen naar nl/_data/tabellen/{1_knopen,2_routes,3_beelden,4_teksten_nl,5_layouts}.json
3. Eventueel: schrijf aparte tekst-tabellen per taal (4_teksten_de.json etc.)

Gebruik:
    python3 scripts/xlsx_naar_json.py

Wordt automatisch uitgevoerd door GitHub Action bij elke push die vizier.xlsx wijzigt.
"""
import json
from pathlib import Path
import sys

try:
    import openpyxl
except ImportError:
    print("Fout: openpyxl niet geinstalleerd. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

REPO = Path(__file__).resolve().parent.parent
XLSX = REPO / "nl" / "_data" / "vizier.xlsx"
TABDIR = REPO / "nl" / "_data" / "tabellen"

if not XLSX.exists():
    print(f"Fout: {XLSX} niet gevonden", file=sys.stderr)
    sys.exit(1)

TABDIR.mkdir(parents=True, exist_ok=True)

wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb["Knopen"]

# Lees header (rij 2)
headers = [cell.value for cell in ws[2]]
# Bouw mapping van header -> kolomindex (0-based)
# Headers beginnen in kolom B (index 1)
header_lookup = {}
for i, h in enumerate(headers):
    if h:
        # Normaliseer headers naar dezelfde keys als JSON-velden
        key = h.lower().replace(" ", "_").replace("-", "_")
        # Specifieke aliassen
        aliassen = {
            "code": "code", "ouder": "ouder", "taal": "taal", "type": "type",
            "naam": "naam", "ondertitel": "ondertitel", "beschrijving": "beschrijving",
            "url": "url", "hero": "hero",
            "hero_positie": "hero_positie", "hero_filter": "hero_filter",
            "tekst_positie": "tekst_positie", "tekst_kleur": "tekst_kleur",
            "terug_naar": "terug_naar", "klik_actie": "klik_actie", "doel_open": "doel_open",
            "status": "status", "volgorde": "volgorde", "actief": "actief",
            "kleur": "kleur", "grad_start": "gradient_start", "grad_eind": "gradient_eind",
            "notitie": "notitie",
            "versie_uitgebreid": "versie_uitgebreid", "versie_kort": "versie_kort",
            "audio_nl": "audio_nl", "audio_de": "audio_de", "audio_en": "audio_en", "audio_ru": "audio_ru",
            "video_url": "video_url", "transcript": "transcript",
            "pdf_download": "pdf_download", "delen_url": "delen_url",
            "tags": "tags", "gerelateerd": "gerelateerd",
            "datum_publicatie": "datum_publicatie", "auteur": "auteur",
            "talen_beschikbaar": "talen_beschikbaar",
            "extra_1": "extra_1", "extra_2": "extra_2", "extra_3": "extra_3",
            "extra_4": "extra_4", "extra_5": "extra_5",
        }
        key = aliassen.get(key, key)
        header_lookup[key] = i

print(f"Headers gevonden: {len(header_lookup)}")

# Lees alle data-rijen (vanaf rij 3)
rijen_ruw = []
for row in ws.iter_rows(min_row=3, values_only=True):
    if not row[header_lookup["code"]]:
        continue
    rij = {}
    for key, idx in header_lookup.items():
        v = row[idx] if idx < len(row) else None
        if v is None:
            v = ""
        elif isinstance(v, bool):
            pass
        elif key == "actief":
            v = str(v).strip().upper() in ("TRUE","1","JA","WAAR")
        else:
            v = str(v).strip()
        rij[key] = v
    rijen_ruw.append(rij)

print(f"Data-rijen gelezen: {len(rijen_ruw)}")

# === Bouw de 5 tabellen ===

# 1. KNOPEN — code, ouder, type, status, volgorde, actief
knopen = {
    "_meta": {
        "tabel": "knopen",
        "versie": "2.0",
        "bron": "vizier.xlsx",
        "beschrijving": "Wat bestaat er en hoe is het gerangschikt. Eerste cijfer van code = taal-ID."
    },
    "rijen": []
}
for r in rijen_ruw:
    if not r.get("actief", True):
        continue
    knopen["rijen"].append({
        "code": r["code"],
        "ouder": r.get("ouder", "") or None,
        "type": r.get("type", "artikel"),
        "status": r.get("status", "live"),
        "volgorde": r.get("volgorde") or None,
        "actief": True,
    })

# 2. ROUTES — url, terug_naar, klik_actie, doel_open + functies (versie_kort, audio, video, etc.)
routes = {
    "_meta": {
        "tabel": "routes",
        "versie": "2.0",
        "bron": "vizier.xlsx",
        "beschrijving": "Wat gebeurt bij klikken. Bevat ook alle alternatieve formaten (kort, audio, video, pdf)."
    },
    "rijen": []
}
for r in rijen_ruw:
    routes["rijen"].append({
        "code": r["code"],
        "url": r.get("url", ""),
        "doel_open": r.get("doel_open", "self") or "self",
        "terug_naar": r.get("terug_naar", ""),
        "klik_actie": r.get("klik_actie", "auto") or "auto",
        "versie_uitgebreid": r.get("versie_uitgebreid", ""),
        "versie_kort": r.get("versie_kort", ""),
        "audio_nl": r.get("audio_nl", ""),
        "audio_de": r.get("audio_de", ""),
        "audio_en": r.get("audio_en", ""),
        "audio_ru": r.get("audio_ru", ""),
        "video_url": r.get("video_url", ""),
        "transcript": r.get("transcript", ""),
        "pdf_download": r.get("pdf_download", ""),
        "delen_url": r.get("delen_url", ""),
    })

# 3. BEELDEN — hero, positie, filter
beelden = {
    "_meta": {
        "tabel": "beelden",
        "versie": "2.0",
        "bron": "vizier.xlsx",
        "filter_definities": {
            "standaard": "sepia(.35) brightness(.85) contrast(1.05)",
            "licht": "sepia(.20) brightness(1.00) contrast(1.05)",
            "donker": "sepia(.55) brightness(.55) contrast(1.10)",
            "helder": "sepia(.10) brightness(1.05) contrast(1.05)",
            "sepia-zwaar": "sepia(.85) brightness(.75) contrast(1.20)",
            "geen": "none"
        }
    },
    "rijen": []
}
for r in rijen_ruw:
    beelden["rijen"].append({
        "code": r["code"],
        "hero": r.get("hero", ""),
        "hero_positie": r.get("hero_positie", "center") or "center",
        "hero_filter": r.get("hero_filter", "standaard") or "standaard",
        "hero_alt": "",
    })

# 4. TEKSTEN — per taal apart bestand
# Groepeer rijen per taal
per_taal = {}
for r in rijen_ruw:
    taal = r.get("taal", "").strip()
    if not taal:
        continue
    per_taal.setdefault(taal, []).append(r)

for taal, rij_lijst in per_taal.items():
    pad = TABDIR / f"4_teksten_{taal}.json"
    # MERGE-MODUS: lees bestaande JSON en behoud rijen die niet in Excel zitten
    bestaande_rijen = {}
    if pad.exists():
        try:
            with open(pad) as f:
                oud = json.load(f)
            for r in oud.get("rijen", []):
                bestaande_rijen[r["code"]] = r
        except Exception:
            pass

    # Overschrijf met Excel-data
    for r in rij_lijst:
        bestaande_rijen[r["code"]] = {
            "code": r["code"],
            "naam": r.get("naam", ""),
            "ondertitel": r.get("ondertitel", ""),
            "beschrijving": r.get("beschrijving", ""),
            "tekst_positie": r.get("tekst_positie", "auto") or "auto",
            "tekst_kleur": r.get("tekst_kleur", "licht") or "licht",
            "notitie": r.get("notitie", ""),
            "tags": r.get("tags", ""),
            "auteur": r.get("auteur", ""),
            "datum_publicatie": r.get("datum_publicatie", ""),
            "gerelateerd": r.get("gerelateerd", ""),
        }

    # Sorteer op code (natuurlijk) en schrijf weg
    def code_sort_key(c):
        return [int(x) if x.isdigit() else x for x in c.split(".")]
    gesorteerd = sorted(bestaande_rijen.values(), key=lambda r: code_sort_key(r["code"]))

    tekst_bestand = {
        "_meta": {
            "tabel": "teksten",
            "versie": "2.0",
            "taal": taal,
            "bron": "vizier.xlsx + bestaande JSON (merge)",
        },
        "rijen": gesorteerd,
    }
    json.dump(tekst_bestand, open(pad, "w"), ensure_ascii=False, indent=2)

# 5. LAYOUTS — uit Layouts-tab
layouts = {
    "_meta": {
        "tabel": "layouts",
        "versie": "2.0",
        "bron": "vizier.xlsx",
    },
    "layouts": {}
}
if "Layouts" in wb.sheetnames:
    ws_l = wb["Layouts"]
    for row in ws_l.iter_rows(min_row=3, values_only=True):
        if not row[1]:  # kolom B = aantal
            continue
        n = str(row[1])
        kolommen = row[2] or "1fr"
        rijen_str = row[3] or "1fr"
        posities_raw = row[4] or ""
        notitie = row[5] or ""
        posities = []
        for stuk in posities_raw.split("|"):
            stuk = stuk.strip()
            if not stuk: continue
            delen = [p.strip() for p in stuk.split(",")]
            if len(delen) >= 2:
                pos = {"col": delen[0], "row": delen[1]}
                if len(delen) >= 3:
                    pos["label"] = delen[2]
                posities.append(pos)
        layouts["layouts"][n] = {
            "kolommen": kolommen,
            "rijen": rijen_str,
            "posities": posities,
        }
        if notitie:
            layouts["layouts"][n]["notitie"] = notitie

# Schrijf alle 4 niet-tekst tabellen
for naam, tabel in [
    ("1_knopen.json", knopen),
    ("2_routes.json", routes),
    ("3_beelden.json", beelden),
    ("5_layouts.json", layouts),
]:
    p = TABDIR / naam
    json.dump(tabel, open(p, "w"), ensure_ascii=False, indent=2)
    print(f"  ✓ {p.name}: {len(tabel.get('rijen', [])) or len(tabel.get('layouts', {}))} rijen")

print(f"\nTekst-tabellen geschreven voor talen: {list(per_taal.keys())}")
print(f"Klaar.")
