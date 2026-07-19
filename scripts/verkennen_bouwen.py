#!/usr/bin/env python3
"""
Vervang de tegel-visualisatie in verkennen.html door een normaal menu.

Genereert per taal een statische lijst van alle artikelen, gegroepeerd per
rubriek. Vervangt de <main class="verkennen-main">...</main> in elk
{taal}/verkennen.html door een schone HTML-lijst.

Bron: nl/_data/vizier.xlsx (Type='artikel', bepaalt per taal).
"""
from __future__ import annotations

import re
import sys
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl niet beschikbaar", file=sys.stderr)
    sys.exit(1)

REPO = Path(__file__).resolve().parent.parent
XLSX = REPO / "nl" / "_data" / "vizier.xlsx"

# ---------------------------------------------------------------------------
# Rubriek-classifier (URL-based)
# ---------------------------------------------------------------------------
EDITIE_MAP = {
    'editie-0': 'Editie Europa', 'edition-0': 'Editie Europa',
    'ausgabe-0': 'Editie Europa', 'vypusk-0': 'Editie Europa',
    'editie-1': 'Editie 1 · Samenhang', 'edition-1': 'Editie 1 · Samenhang',
    'ausgabe-1': 'Editie 1 · Samenhang', 'vypusk-1': 'Editie 1 · Samenhang',
    'editie-2': 'Editie 2 · Belastingbeleid', 'edition-2': 'Editie 2 · Belastingbeleid',
    'ausgabe-2': 'Editie 2 · Belastingbeleid', 'vypusk-2': 'Editie 2 · Belastingbeleid',
    'editie-3': 'Editie 3 · Onder het ijs', 'edition-3': 'Editie 3 · Onder het ijs',
    'ausgabe-3': 'Editie 3 · Onder het ijs', 'vypusk-3': 'Editie 3 · Onder het ijs',
    'editie-4': 'Editie 4 · Papier-macht', 'edition-4': 'Editie 4 · Papier-macht',
    'ausgabe-4': 'Editie 4 · Papier-macht', 'vypusk-4': 'Editie 4 · Papier-macht',
    'editie-5': 'Editie 5 · Vergeten orde', 'edition-5': 'Editie 5 · Vergeten orde',
    'ausgabe-5': 'Editie 5 · Vergeten orde', 'vypusk-5': 'Editie 5 · Vergeten orde',
    'editie-6': 'Editie 6 · Nova Democratia', 'edition-6': 'Editie 6 · Nova Democratia',
    'ausgabe-6': 'Editie 6 · Nova Democratia', 'vypusk-6': 'Editie 6 · Nova Democratia',
    'editie-klimaat': 'Editie Klimaat', 'climate-edition': 'Editie Klimaat',
    'ausgabe-klima': 'Editie Klimaat',
    'editie-stikstof': 'Editie Stikstof', 'nitrogen-edition': 'Editie Stikstof',
    'ausgabe-stickstoff': 'Editie Stikstof',
    'editie-duitsland': 'Editie Duitsland', 'germany-edition': 'Editie Duitsland',
    'ausgabe-deutschland': 'Editie Duitsland',
    'editie-leiderschap': 'Editie Leiderschap',
    'leadership-edition': 'Editie Leiderschap',
    'ausgabe-fuehrung': 'Editie Leiderschap',
}

RUBRIEK_VOLGORDE = [
    "Wat opkomt",
    "Editie Europa",
    "Editie 1 · Samenhang",
    "Editie 2 · Belastingbeleid",
    "Editie 3 · Onder het ijs",
    "Editie 4 · Papier-macht",
    "Editie 5 · Vergeten orde",
    "Editie 6 · Nova Democratia",
    "Editie Klimaat",
    "Editie Stikstof",
    "Editie Duitsland",
    "Editie Leiderschap",
    "Gevolgenkaarten",
    "Onderzoek",
    "Dossiers",
    "Filosofie",
]

# Rubriek-labels per taal
RUBRIEK_LABELS = {
    'nl': {
        "Wat opkomt": "Nieuw",
        "Gevolgenkaarten": "Gevolgenkaarten",
        "Onderzoek": "Onderzoek",
        "Dossiers": "Dossiers",
        "Filosofie": "Filosofie",
    },
    'en': {
        "Wat opkomt": "Latest",
        "Editie Europa": "Europe Edition",
        "Editie 1 · Samenhang": "Edition 1 · Coherence",
        "Editie 2 · Belastingbeleid": "Edition 2 · Tax policy",
        "Editie 3 · Onder het ijs": "Edition 3 · Under the ice",
        "Editie 4 · Papier-macht": "Edition 4 · Paper power",
        "Editie 5 · Vergeten orde": "Edition 5 · Forgotten order",
        "Editie 6 · Nova Democratia": "Edition 6 · Nova Democratia",
        "Editie Klimaat": "Climate Edition",
        "Editie Stikstof": "Nitrogen Edition",
        "Editie Duitsland": "Germany Edition",
        "Editie Leiderschap": "Leadership Edition",
        "Gevolgenkaarten": "Consequence maps",
        "Onderzoek": "Research",
        "Dossiers": "Dossiers",
        "Filosofie": "Philosophy",
    },
    'de': {
        "Wat opkomt": "Neueste",
        "Editie Europa": "Ausgabe Europa",
        "Editie 1 · Samenhang": "Ausgabe 1 · Zusammenhang",
        "Editie 2 · Belastingbeleid": "Ausgabe 2 · Steuerpolitik",
        "Editie 3 · Onder het ijs": "Ausgabe 3 · Unter dem Eis",
        "Editie 4 · Papier-macht": "Ausgabe 4 · Papier-Macht",
        "Editie 5 · Vergeten orde": "Ausgabe 5 · Vergessene Ordnung",
        "Editie 6 · Nova Democratia": "Ausgabe 6 · Nova Democratia",
        "Editie Klimaat": "Ausgabe Klima",
        "Editie Stikstof": "Ausgabe Stickstoff",
        "Editie Duitsland": "Ausgabe Deutschland",
        "Editie Leiderschap": "Ausgabe Führung",
        "Gevolgenkaarten": "Folgenkarten",
        "Onderzoek": "Forschung",
        "Dossiers": "Dossiers",
        "Filosofie": "Philosophie",
    },
    'ru': {
        "Wat opkomt": "Новое",
        "Editie Europa": "Выпуск Европа",
        "Editie 1 · Samenhang": "Выпуск 1 · Согласованность",
        "Editie 2 · Belastingbeleid": "Выпуск 2 · Налоговая политика",
        "Editie 3 · Onder het ijs": "Выпуск 3 · Подо льдом",
        "Editie 4 · Papier-macht": "Выпуск 4 · Власть бумаги",
        "Editie 5 · Vergeten orde": "Выпуск 5 · Забытый порядок",
        "Editie 6 · Nova Democratia": "Выпуск 6 · Nova Democratia",
        "Editie Klimaat": "Выпуск Климат",
        "Editie Stikstof": "Выпуск Азот",
        "Editie Duitsland": "Выпуск Германия",
        "Editie Leiderschap": "Выпуск Лидерство",
        "Gevolgenkaarten": "Карты последствий",
        "Onderzoek": "Исследования",
        "Dossiers": "Досье",
        "Filosofie": "Философия",
    },
}


def classify(url: str) -> str:
    u = (url or '').lower()
    if 'wat-opkomt' in u or 'what-surfaces' in u or 'was-aufkommt' in u or 'chto-vsplyvaet' in u:
        return "Wat opkomt"
    if 'dossier' in u:
        return "Dossiers"
    if 'gevolgenkaart' in u or 'consequence' in u or 'folgenkarte' in u:
        return "Gevolgenkaarten"
    if 'onderzoek' in u or 'research' in u or 'forschung' in u or 'issledovanie' in u:
        return "Onderzoek"
    if 'filosofie' in u or 'philosophy' in u or 'philosophie' in u:
        return "Filosofie"
    for key, label in EDITIE_MAP.items():
        if key in u:
            return label
    return "Overig"


# ---------------------------------------------------------------------------
# Data laden
# ---------------------------------------------------------------------------
def load_articles():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb['Knopen']
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=2, column=c).value
        if v:
            headers[v.strip()] = c
    C = headers

    per_taal = defaultdict(lambda: defaultdict(list))  # taal → rubriek → [artikelen]
    for r in range(3, ws.max_row + 1):
        t = ws.cell(row=r, column=C['Type']).value
        l = ws.cell(row=r, column=C['Taal']).value
        if not (t and l):
            continue
        if str(t).strip() != 'artikel':
            continue
        lang = str(l).strip()
        if lang not in ('nl', 'en', 'de', 'ru'):
            continue
        url = ws.cell(row=r, column=C['URL']).value or ''
        naam = ws.cell(row=r, column=C['Naam']).value or ''
        ondertitel = ws.cell(row=r, column=C['Ondertitel']).value or ''
        beschrijving = ws.cell(row=r, column=C['Beschrijving']).value or ''
        status = ws.cell(row=r, column=C['Status']).value or ''
        actief = ws.cell(row=r, column=C['Actief']).value
        actief = True if actief is None else bool(actief)
        volgorde = ws.cell(row=r, column=C['Volgorde']).value

        if not str(url).strip():
            continue  # skip artikelen zonder URL
        if str(status).strip().lower() == 'concept':
            continue
        if not actief:
            continue

        rubriek = classify(str(url))
        korte = (str(ondertitel).strip() or str(beschrijving).strip())[:200]
        per_taal[lang][rubriek].append({
            'naam': str(naam).strip(),
            'korte': korte,
            'url': str(url).strip(),
            'volgorde': volgorde if isinstance(volgorde, (int, float)) else 999,
        })
    return per_taal


# ---------------------------------------------------------------------------
# HTML renderen
# ---------------------------------------------------------------------------
def render_menu_html(taal: str, per_rubriek: dict) -> str:
    labels = RUBRIEK_LABELS.get(taal, {})
    intro_labels = {
        'nl': ("Overzicht", "Alle artikelen van Het Open Vizier, gegroepeerd per rubriek. Klik op een titel om het artikel te openen."),
        'en': ("Overview", "All articles from Het Open Vizier, grouped by section. Click a title to open the article."),
        'de': ("Übersicht", "Alle Artikel von Het Open Vizier, nach Rubrik gruppiert. Klicken Sie auf einen Titel, um den Artikel zu öffnen."),
        'ru': ("Обзор", "Все статьи «Het Open Vizier», сгруппированные по рубрикам. Нажмите на заголовок, чтобы открыть статью."),
    }
    titel, sub = intro_labels.get(taal, intro_labels['nl'])

    parts = ['<main class="verkennen-main">']
    parts.append(f'  <section class="verkennen-intro">')
    parts.append(f'    <h1>{titel}</h1>')
    parts.append(f'    <p>{sub}</p>')
    parts.append(f'  </section>')
    parts.append('  <div class="vk-menu">')

    for rub in RUBRIEK_VOLGORDE:
        if rub not in per_rubriek:
            continue
        artikelen = per_rubriek[rub]
        if not artikelen:
            continue
        artikelen.sort(key=lambda a: (float(a['volgorde']) if isinstance(a['volgorde'], (int, float)) else 999, a['naam'].lower()))
        label = labels.get(rub, rub)
        parts.append(f'    <section class="vk-rub">')
        parts.append(f'      <h2 class="vk-rub__titel">{label}</h2>')
        parts.append(f'      <ul class="vk-rub__lijst">')
        for a in artikelen:
            desc = a['korte']
            desc_html = f'<span class="vk-item__desc">{desc}</span>' if desc else ''
            parts.append(f'        <li class="vk-item"><a href="{a["url"]}"><span class="vk-item__titel">{a["naam"]}</span>{desc_html}</a></li>')
        parts.append(f'      </ul>')
        parts.append(f'    </section>')

    parts.append('  </div>')
    parts.append('</main>')
    return "\n".join(parts)


VK_CSS = """<style>
  /* verkennen-menu v1 — platte lijst per rubriek */
  .verkennen-main { max-width:1100px; margin:0 auto; padding:2rem 1.25rem 4rem; }
  .verkennen-intro { text-align:center; margin-bottom:2.5rem; }
  .verkennen-intro h1 { font-family:Georgia,serif; font-size:clamp(2rem,3.5vw,2.8rem); font-weight:400; color:#1a1a1a; margin-bottom:0.75rem; }
  .verkennen-intro p { color:#6b7280; font-size:1.05rem; line-height:1.5; max-width:640px; margin:0 auto; }
  .vk-menu { display:flex; flex-direction:column; gap:2.5rem; }
  .vk-rub__titel { font-family:Georgia,serif; font-size:1.35rem; color:#7a1f2b; font-weight:600; padding-bottom:0.5rem; border-bottom:2px solid #d4af37; margin-bottom:1rem; }
  .vk-rub__lijst { list-style:none; padding:0; margin:0; }
  .vk-item { margin:0; padding:0; border-bottom:1px solid #efece3; }
  .vk-item:last-child { border-bottom:0; }
  .vk-item a { display:block; padding:0.9rem 0.5rem; color:#1a1a1a; text-decoration:none; transition:background 0.15s; }
  .vk-item a:hover { background:#f5f3ee; padding-left:1rem; }
  .vk-item__titel { display:block; font-weight:600; font-size:1.05rem; color:#1a1a1a; line-height:1.35; }
  .vk-item__desc { display:block; margin-top:0.25rem; font-size:0.92rem; color:#6b7280; line-height:1.45; }
  @media (max-width:640px) {
    .verkennen-main { padding:1.25rem 0.75rem 3rem; }
    .vk-menu { gap:2rem; }
    .vk-rub__titel { font-size:1.2rem; }
  }
</style>
"""


# ---------------------------------------------------------------------------
# Injectie in verkennen.html
# ---------------------------------------------------------------------------
def inject(html: str, new_main: str) -> str:
    # Vervang <main class="verkennen-main"> ... </main> volledig
    new_html, count = re.subn(
        r'<main class="verkennen-main"[^>]*>.*?</main>',
        lambda m: new_main,
        html, count=1, flags=re.S,
    )
    if count == 0:
        raise RuntimeError("Geen <main class=\"verkennen-main\"> gevonden")

    # Verwijder oude embedded <style> voor .verkennen-canvas etc. (blijven of niet? — laten staan)
    # Verwijder eventuele oude vk-menu v1 css uit vorige run
    new_html = re.sub(r'<style>\s*\n?\s*/\* verkennen-menu v1.*?</style>\s*', '', new_html, flags=re.S)
    # Voeg VK_CSS toe vóór </head>
    new_html = new_html.replace('</head>', VK_CSS + '</head>', 1)

    # Verwijder de grote tegel-JavaScript (script-blok dat KNOPEN/CHILDREN/gridLayout etc. gebruikt)
    # Alleen als het aanwezig is
    new_html = re.sub(
        r'<script>\s*\nlet KNOPEN = \[\].*?</script>',
        '',
        new_html, count=1, flags=re.S,
    )
    return new_html


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    print("Artikelen laden…")
    data = load_articles()
    for taal in ('nl', 'en', 'de', 'ru'):
        per_rub = data.get(taal, {})
        print(f"\n[{taal}] {sum(len(v) for v in per_rub.values())} artikelen in {len(per_rub)} rubrieken")
        for rub, arts in per_rub.items():
            print(f"   {rub}: {len(arts)}")

        path = REPO / taal / "verkennen.html"
        if not path.exists():
            print(f"   ! {path.relative_to(REPO)} bestaat niet")
            continue
        html = path.read_text(encoding='utf-8')
        new_main = render_menu_html(taal, per_rub)
        try:
            new_html = inject(html, new_main)
        except RuntimeError as e:
            print(f"   ! {e}")
            continue
        path.write_text(new_html, encoding='utf-8')
        print(f"   ✓ {path.relative_to(REPO)} bijgewerkt ({len(new_html)} bytes)")

    print("\nKlaar.")


if __name__ == "__main__":
    main()
