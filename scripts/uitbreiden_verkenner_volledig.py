#!/usr/bin/env python3
"""
Vul de drie gaten in de Verkenner-structuur, voor alle 4 hoofdtalen (NL=1, DE=2, EN=3, RU=4):

GAT 1: '<wortel>.1.1' Laatste artikelen — slechts 1 artikel (voedingslijn), uitbreiden tot 16
GAT 2: '<wortel>.3.1' Dossiers — 0 kinderen, toevoegen: Vortex Hair
GAT 3: '<wortel>.3.2' Onderzoek — 0 kinderen, toevoegen: EU-multiplicator + Ontwikkelingslandstatus

Werkwijze:
- Voeg rijen toe aan nl/_data/vizier.xlsx (één Excel, taalafhankelijke kolommen per knoop)
- Update 2_routes_<lang>.json per taal met de URL-routes
- Wortel-prefix per taal: nl=1, de=2, en=3, ru=4
"""
from pathlib import Path
import json
from openpyxl import load_workbook
from copy import copy

REPO = Path("/tmp/gh-repo")
XLSX = REPO / "nl/_data/vizier.xlsx"
TAB = REPO / "nl/_data/tabellen"

# Per taal: prefix-cijfer in code, en de wat-opkomt/dossiers/onderzoek paden
LANGS = {
    "nl": {"prefix": "1", "wo": "wat-opkomt",     "doss": "dossiers", "ond": "onderzoek"},
    "de": {"prefix": "2", "wo": "was-aufkommt",   "doss": "dossiers", "ond": "forschung"},
    "en": {"prefix": "3", "wo": "what-surfaces",  "doss": "dossiers", "ond": "research"},
    "ru": {"prefix": "4", "wo": "chto-vsplyvaet", "doss": "dossiers", "ond": "issledovanie"},
}

# === LAATSTE ARTIKELEN (15 nieuwe + 1 bestaande voedingslijn op pos 6) ===
# Tuple: (subcode_onder_X.1.1, slug_per_taal_dict, naam_per_taal, ondertitel_per_taal)
# Voor de eenvoud houden we voor DE/EN/RU dezelfde Engelse-of-vertaalde structuur — de naam wordt
# uit het bestand zelf gehaald (titels zijn al vertaald in de artikelen).
LAATSTE_ARTIKELEN = [
    # (subcode, slugs_per_lang, naam_per_lang, ondertitel_per_lang)
    ("1", {  # 1.1.1.1 reeds aanwezig (voedingslijn-NL) — wij voegen vertalingen toe en bewaren NL ondertitel
        "nl": "de-methodologie-van-de-aanpassing",
        # voor andere talen: route-only, geen XLSX-rij want 1.1.1.1 bestaat al
    }, {
        "nl": "De voedingslijn",
    }, {
        "nl": "Hoe klimaatoverheid, linkse partijen en vakbonden dezelfde fout maken",
    }),
    ("2", {
        "nl": "de-metafoor-auto-immuunziekte",
        "de": "die-metapher-autoimmunkrankheit",
        "en": "the-metaphor-autoimmune-disease",
        "ru": "metafora-autoimmunnoy-bolezni",
    }, {
        "nl": "Een staat met een auto-immuunziekte",
        "de": "Ein Staat mit einer Autoimmunkrankheit",
        "en": "A state with an autoimmune disease",
        "ru": "Государство с аутоиммунной болезнью",
    }, {
        "nl": "Een beeld om mee te denken — de metafoor die de Brussel-stukken draagt.",
        "de": "Ein Bild zum Mitdenken — die Metapher, die die Brüssel-Stücke trägt.",
        "en": "An image to think with — the metaphor that carries the Brussels pieces.",
        "ru": "Образ для размышления — метафора, на которой держатся брюссельские материалы.",
    }),
    ("3", {
        "nl": "de-anti-immuunziekte-van-brussel",
        "de": "die-anti-immunkrankheit-bruessels",
        "en": "the-anti-immune-disease-of-brussels",
        "ru": "anti-immunnaya-bolezn-bryusselya",
    }, {
        "nl": "De anti-immuunziekte van Brussel maakt dat we arm worden",
        "de": "Die Anti-Immunkrankheit Brüssels macht uns arm",
        "en": "The anti-immune disease of Brussels is making us poor",
        "ru": "Анти-иммунная болезнь Брюсселя делает нас бедными",
    }, {
        "nl": "CBAM, ETS en Pillar Two — drie aanvallen tegelijk. De BiCRS-route wordt genegeerd.",
        "de": "CBAM, ETS und Pillar Two — drei Angriffe gleichzeitig. Die BiCRS-Route wird ignoriert.",
        "en": "CBAM, ETS and Pillar Two — three attacks at once. The BiCRS route is ignored.",
        "ru": "CBAM, ETS и Pillar Two — три удара одновременно. Маршрут BiCRS игнорируется.",
    }),
    ("4", {
        "nl": "zij-doden-hun-levensaders-brussel",
        "de": "sie-toeten-ihre-lebensadern-bruessel",
        "en": "they-kill-their-lifelines-brussels",
        "ru": "oni-ubivayut-svoi-zhiznennye-arterii-bryussel",
    }, {
        "nl": "Zij doden hun levensaders",
        "de": "Sie töten ihre Lebensadern",
        "en": "They kill their lifelines",
        "ru": "Они убивают свои жизненные артерии",
    }, {
        "nl": "De diagnose in zes sectoren: één leiding, één anti-immuunziekte.",
        "de": "Die Diagnose in sechs Sektoren: eine Führung, eine Anti-Immunkrankheit.",
        "en": "The diagnosis across six sectors: one leadership, one anti-immune disease.",
        "ru": "Диагноз по шести секторам: одно руководство, одна анти-иммунная болезнь.",
    }),
    ("5", {
        "nl": "de-auto-immuunziekte-slaat-toe",
        "de": "die-autoimmunkrankheit-schlaegt-zu",
        "en": "the-autoimmune-disease-strikes",
        "ru": "autoimmunnaya-bolezn-nanosit-udar",
    }, {
        "nl": "De anti-immuunziekte van onze overheid slaat toe",
        "de": "Die Autoimmunkrankheit unserer Regierung schlägt zu",
        "en": "The autoimmune disease of our government strikes",
        "ru": "Аутоиммунная болезнь нашего правительства наносит удар",
    }, {
        "nl": "Den Haag, stikstof — de spiegel van Brussel in eigen tuin.",
        "de": "Den Haag, Stickstoff — der Spiegel Brüssels im eigenen Garten.",
        "en": "The Hague, nitrogen — the mirror of Brussels in its own garden.",
        "ru": "Гаага, азот — зеркало Брюсселя в собственном саду.",
    }),
    ("6", {
        "nl": "zij-doden-hun-levensader",
        "de": "sie-toeten-ihre-lebensader",
        "en": "they-kill-their-lifeline",
        "ru": "oni-ubivayut-svoyu-zhiznennuyu-arteriyu",
    }, {
        "nl": "Zij doden hun levensader",
        "de": "Sie töten ihre Lebensader",
        "en": "They kill their lifeline",
        "ru": "Они убивают свою жизненную артерию",
    }, {
        "nl": "Den Haag-diagnose: boeren en MKB. €212.795 per kilogram stikstof.",
        "de": "Den Haag-Diagnose: Bauern und KMU. 212.795 € pro Kilogramm Stickstoff.",
        "en": "The Hague diagnosis: farmers and SMEs. €212,795 per kilogram of nitrogen.",
        "ru": "Гаагский диагноз: фермеры и МСП. 212 795 € за килограмм азота.",
    }),
    ("7", {
        "nl": "het-immuunsysteem-van-organisaties",
        # andere talen: nog niet vertaald, route wijst naar NL
    }, {
        "nl": "Het immuunsysteem dat gezonde cellen aanvalt",
    }, {
        "nl": "Opinie — de medische metafoor toegepast op de Nederlandse staat.",
    }),
    ("8", {
        "nl": "de-grote-plundering",
        "de": "die-grosse-pluenderung",
        "en": "the-great-plunder",
        "ru": "velikoe-razgrablenie",
    }, {
        "nl": "De Grote Plundering",
        "de": "Die Große Plünderung",
        "en": "The Great Plunder",
        "ru": "Великое разграбление",
    }, {
        "nl": "Vijfluik over hoe Europa zijn productieven leegrooft.",
        "de": "Fünfteiler darüber, wie Europa seine Produktiven ausräumt.",
        "en": "A five-part series on how Europe is stripping its productive citizens.",
        "ru": "Пятичастная серия о том, как Европа разоряет своих производителей.",
    }),
    ("9", {
        "nl": "plundering-1-diagnose",
        "de": "pluenderung-1-diagnose",
        "en": "plunder-1-diagnosis",
        "ru": "razgrablenie-1-diagnoz",
    }, {
        "nl": "Plundering I — Diagnose",
        "de": "Plünderung I — Diagnose",
        "en": "Plunder I — Diagnosis",
        "ru": "Разграбление I — Диагноз",
    }, {
        "nl": "Deel 1: wij rupfen wie ons voedt.",
        "de": "Teil 1: Wir rupfen, wer uns nährt.",
        "en": "Part 1: we pluck whoever feeds us.",
        "ru": "Часть 1: мы общипываем тех, кто нас кормит.",
    }),
    ("10", {
        "nl": "plundering-2-mechaniek",
        "de": "pluenderung-2-mechanik",
        "en": "plunder-2-mechanics",
        "ru": "razgrablenie-2-mekhanika",
    }, {
        "nl": "Plundering II — Mechaniek",
        "de": "Plünderung II — Mechanik",
        "en": "Plunder II — Mechanics",
        "ru": "Разграбление II — Механика",
    }, {
        "nl": "Deel 2: hoe het systeem zijn eigen producenten ontmoedigt.",
        "de": "Teil 2: Wie das System seine eigenen Produzenten entmutigt.",
        "en": "Part 2: how the system discourages its own producers.",
        "ru": "Часть 2: как система разубеждает собственных производителей.",
    }),
    ("11", {
        "nl": "plundering-3-afloop",
        "de": "pluenderung-3-ausgang",
        "en": "plunder-3-outcome",
        "ru": "razgrablenie-3-iskhod",
    }, {
        "nl": "Plundering III — Afloop",
        "de": "Plünderung III — Ausgang",
        "en": "Plunder III — Outcome",
        "ru": "Разграбление III — Исход",
    }, {
        "nl": "Deel 3: wat resteert wanneer de productieven vertrokken zijn.",
        "de": "Teil 3: Was übrig bleibt, wenn die Produktiven gegangen sind.",
        "en": "Part 3: what remains once the productive citizens have left.",
        "ru": "Часть 3: что остаётся, когда производители уходят.",
    }),
    ("12", {
        "nl": "plundering-4-politieke-landschap",
        "de": "pluenderung-4-politische-landschaft",
        "en": "plunder-4-political-landscape",
        "ru": "razgrablenie-4-politicheskij-landshaft",
    }, {
        "nl": "Plundering IV — Politieke landschap",
        "de": "Plünderung IV — Politische Landschaft",
        "en": "Plunder IV — Political landscape",
        "ru": "Разграбление IV — Политический ландшафт",
    }, {
        "nl": "Deel 4: hoe partij na partij meegaat in dezelfde fout.",
        "de": "Teil 4: Wie eine Partei nach der anderen den gleichen Fehler macht.",
        "en": "Part 4: how party after party falls for the same mistake.",
        "ru": "Часть 4: как партия за партией совершает ту же ошибку.",
    }),
    ("13", {
        "nl": "plundering-0-eerst-plukken-dan-oordelen",
        "de": "pluenderung-0-erst-rupfen-dann-urteilen",
        "en": "plunder-0-pluck-first-judge-later",
        "ru": "razgrablenie-0-snachala-obshchipat-potom-sudit",
    }, {
        "nl": "Plundering 0 — Eerst plukken, dan oordelen",
        "de": "Plünderung 0 — Erst rupfen, dann urteilen",
        "en": "Plunder 0 — Pluck first, judge later",
        "ru": "Разграбление 0 — Сначала ощипать, потом судить",
    }, {
        "nl": "Proloog op de tetralogie: de logica van het oordeel-eerst.",
        "de": "Prolog der Tetralogie: die Logik des Urteils-zuerst.",
        "en": "Prologue to the tetralogy: the logic of judgement-first.",
        "ru": "Пролог тетралогии: логика «сначала суд».",
    }),
    ("14", {
        "nl": "de-actor-is-de-regel",
        "de": "der-akteur-ist-die-regel",
        "en": "the-actor-is-the-rule",
        "ru": "aktor-est-pravilo",
    }, {
        "nl": "De actor is de regel",
        "de": "Der Akteur ist die Regel",
        "en": "The actor is the rule",
        "ru": "Актор есть правило",
    }, {
        "nl": "Sluitstuk van de tetralogie — vier lagen van uitdrijving.",
        "de": "Schlussstück der Tetralogie — vier Schichten der Verdrängung.",
        "en": "Closing piece of the tetralogy — four layers of expulsion.",
        "ru": "Заключение тетралогии — четыре слоя вытеснения.",
    }),
    ("15", {
        "nl": "de-singaporese-gevolgenkaart",
        "de": "die-singapurer-folgenkarte",
        "en": "the-singaporean-consequence-map",
        "ru": "singapurskaya-karta-posledstvij",
    }, {
        "nl": "De Singaporese Gevolgenkaart",
        "de": "Die Singapurer Folgenkarte",
        "en": "The Singaporean Consequence Map",
        "ru": "Сингапурская карта последствий",
    }, {
        "nl": "Wat Singapore wel goed doet en Brussel niet.",
        "de": "Was Singapur richtig macht und Brüssel nicht.",
        "en": "What Singapore gets right and Brussels does not.",
        "ru": "Что Сингапур делает правильно, а Брюссель нет.",
    }),
    ("16", {
        "nl": "de-zwitserse-gevolgenkaart",
        "de": "die-schweizer-folgenkarte",
        "en": "the-swiss-consequence-map",
        "ru": "shvejcarskaya-karta-posledstvij",
    }, {
        "nl": "De Zwitserse Gevolgenkaart",
        "de": "Die Schweizer Folgenkarte",
        "en": "The Swiss Consequence Map",
        "ru": "Швейцарская карта последствий",
    }, {
        "nl": "Wat Zwitserland doet dat Brussel niet kan.",
        "de": "Was die Schweiz tut, was Brüssel nicht kann.",
        "en": "What Switzerland does that Brussels cannot.",
        "ru": "Что делает Швейцария, чего не может Брюссель.",
    }),
]

# === DOSSIERS (onder X.3.1) ===
DOSSIERS = [
    ("1", {
        "nl": "vortex-hair", "de": "vortex-hair", "en": "vortex-hair", "ru": "vortex-hair",
    }, {
        "nl": "Vortex Hair",
        "de": "Vortex Hair",
        "en": "Vortex Hair",
        "ru": "Vortex Hair",
    }, {
        "nl": "Een dossier over de wervelinggetekende haarverzorging.",
        "de": "Ein Dossier über die wirbelmarkierte Haarpflege.",
        "en": "A dossier on vortex-marked hair care.",
        "ru": "Досье о вихревом уходе за волосами.",
    }),
]

# === ONDERZOEK (onder X.3.2) ===
ONDERZOEK = [
    ("1", {
        "nl": "eu-multiplicator",
        "de": "eu-multiplikator",
        "en": "eu-multiplier",
        "ru": "es-multiplikator",
    }, {
        "nl": "De EU-multiplicator",
        "de": "Der EU-Multiplikator",
        "en": "The EU multiplier",
        "ru": "ЕС-мультипликатор",
    }, {
        "nl": "Onderzoek naar de versterkings­factor van Brusselse maatregelen.",
        "de": "Untersuchung des Verstärkungs­faktors Brüsseler Maßnahmen.",
        "en": "Research on the multiplication factor of Brussels-level policies.",
        "ru": "Исследование коэффициента усиления брюссельских мер.",
    }),
    ("2", {
        "nl": "ontwikkelingslandstatus",
        "de": "entwicklungslandstatus",
        "en": "developing-country-status",
        "ru": "status-razvivayushcheysya-strany",
    }, {
        "nl": "Ontwikkelingslandstatus",
        "de": "Entwicklungslandstatus",
        "en": "Developing country status",
        "ru": "Статус развивающейся страны",
    }, {
        "nl": "Onderzoek naar de mogelijke route via WTO/IMF.",
        "de": "Untersuchung der möglichen Route über WTO/IWF.",
        "en": "Research on the possible route via WTO/IMF.",
        "ru": "Исследование возможного пути через ВТО/МВФ.",
    }),
]

# =============================================================================
# Stap 1: XLSX bijwerken
# =============================================================================
wb = load_workbook(XLSX)
ws = wb['Knopen']

headers = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column+1)]
def col(name):
    return headers.index(name) + 1 if name in headers else None

c_code   = col("Code")
c_ouder  = col("Ouder")
c_taal   = col("Taal")
c_type   = col("Type")
c_naam   = col("Naam")
c_ondt   = col("Ondertitel")
c_herofilter = col("Hero-filter")
c_status = col("Status")
c_volg   = col("Volgorde")
c_actief = col("Actief")
c_auteur = col("Auteur")

# Verzamel bestaande codes
existing_codes = set()
last_row = 2
for r in range(3, ws.max_row + 1):
    code = ws.cell(r, c_code).value
    if code:
        existing_codes.add(str(code))
        last_row = r

def add_row(code, ouder, taal, naam, ondt, volg):
    """Voeg één Excel-rij toe."""
    global last_row
    if code in existing_codes:
        return False
    last_row += 1
    ws.cell(last_row, c_code).value = code
    ws.cell(last_row, c_ouder).value = ouder
    ws.cell(last_row, c_taal).value = taal
    ws.cell(last_row, c_type).value = "artikel"
    ws.cell(last_row, c_naam).value = naam
    ws.cell(last_row, c_ondt).value = ondt
    ws.cell(last_row, c_herofilter).value = "standaard"
    ws.cell(last_row, c_status).value = "live"
    ws.cell(last_row, c_volg).value = volg
    ws.cell(last_row, c_actief).value = True
    ws.cell(last_row, c_auteur).value = "Jacobus van Merksteijn"
    existing_codes.add(code)
    return True

added_xlsx = 0
# Laatste artikelen — voor alle 4 talen
for lang, info in LANGS.items():
    prefix = info["prefix"]
    ouder = f"{prefix}.1.1"
    for subcode, slugs, namen, ondts in LAATSTE_ARTIKELEN:
        if lang not in namen:
            continue
        code = f"{prefix}.1.1.{subcode}"
        if add_row(code, ouder, lang, namen[lang], ondts.get(lang, ""), int(subcode)):
            added_xlsx += 1

# Dossiers
for lang, info in LANGS.items():
    prefix = info["prefix"]
    ouder = f"{prefix}.3.1"
    for subcode, slugs, namen, ondts in DOSSIERS:
        if lang not in namen:
            continue
        code = f"{prefix}.3.1.{subcode}"
        if add_row(code, ouder, lang, namen[lang], ondts.get(lang, ""), int(subcode)):
            added_xlsx += 1

# Onderzoek
for lang, info in LANGS.items():
    prefix = info["prefix"]
    ouder = f"{prefix}.3.2"
    for subcode, slugs, namen, ondts in ONDERZOEK:
        if lang not in namen:
            continue
        code = f"{prefix}.3.2.{subcode}"
        if add_row(code, ouder, lang, namen[lang], ondts.get(lang, ""), int(subcode)):
            added_xlsx += 1

wb.save(XLSX)
print(f"XLSX: {added_xlsx} nieuwe rijen toegevoegd")

# =============================================================================
# Stap 2: 2_routes_<lang>.json bijwerken
# =============================================================================
def add_route(routes_path, code, url):
    routes = json.loads(routes_path.read_text(encoding="utf-8"))
    existing = {r["code"]: r for r in routes["rijen"]}
    if code in existing:
        if existing[code].get("url") != url:
            existing[code]["url"] = url
            routes_path.write_text(json.dumps(routes, ensure_ascii=False, indent=2), encoding="utf-8")
            return "updated"
        return "unchanged"
    # Nieuwe rij
    new_row = {k: None for k in routes["rijen"][0].keys()} if routes["rijen"] else {"code": None, "url": None}
    new_row["code"] = code
    new_row["url"] = url
    new_row["doel_open"] = "zelf"
    routes["rijen"].append(new_row)
    routes_path.write_text(json.dumps(routes, ensure_ascii=False, indent=2), encoding="utf-8")
    return "added"

route_stats = {"added": 0, "updated": 0, "unchanged": 0, "geskipt_geen_bestand": 0}

for lang, info in LANGS.items():
    prefix = info["prefix"]
    routes_path = TAB / f"2_routes_{lang}.json"

    # Laatste artikelen
    for subcode, slugs, namen, ondts in LAATSTE_ARTIKELEN:
        if lang not in slugs:
            continue
        slug = slugs[lang]
        code = f"{prefix}.1.1.{subcode}"
        url = f"{info['wo']}/{slug}.html"
        # Verifieer bestand
        if not (REPO / lang / url).exists():
            route_stats["geskipt_geen_bestand"] += 1
            print(f"  SKIP: {lang}/{url} bestaat niet (code {code})")
            continue
        result = add_route(routes_path, code, url)
        route_stats[result] += 1

    # Dossiers
    for subcode, slugs, namen, ondts in DOSSIERS:
        if lang not in slugs:
            continue
        slug = slugs[lang]
        code = f"{prefix}.3.1.{subcode}"
        url = f"{info['doss']}/{slug}.html"
        if not (REPO / lang / url).exists():
            route_stats["geskipt_geen_bestand"] += 1
            print(f"  SKIP: {lang}/{url} bestaat niet (code {code})")
            continue
        result = add_route(routes_path, code, url)
        route_stats[result] += 1

    # Onderzoek
    for subcode, slugs, namen, ondts in ONDERZOEK:
        if lang not in slugs:
            continue
        slug = slugs[lang]
        code = f"{prefix}.3.2.{subcode}"
        url = f"{info['ond']}/{slug}.html"
        if not (REPO / lang / url).exists():
            route_stats["geskipt_geen_bestand"] += 1
            print(f"  SKIP: {lang}/{url} bestaat niet (code {code})")
            continue
        result = add_route(routes_path, code, url)
        route_stats[result] += 1

print(f"\nRoutes: {route_stats}")
