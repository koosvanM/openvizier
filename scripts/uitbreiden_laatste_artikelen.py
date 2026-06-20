#!/usr/bin/env python3
"""
Voeg de 15 nieuwste NL-artikelen toe aan vizier.xlsx onder ouder 1.1.1
(Bladeren → Laatste artikelen), én voeg URLs toe aan 2_routes_nl.json.

Volgorde-doel (in de Verkenner verschijnt 1 als eerste tegel):
1.  De metafoor — auto-immuunziekte           (proloog, zacht)
2.  De anti-immuunziekte van Brussel          (lead Brussel)
3.  Zij doden hun levensaders                 (diagnose Brussel)
4.  De anti-immuunziekte van onze overheid    (lead Den Haag)
5.  Zij doden hun levensader                  (diagnose Den Haag)
6.  De voedingslijn                            (bestaand, code 1.1.1.1)
7.  Het immuunsysteem dat gezonde cellen aanvalt
8.  De Grote Plundering                       (vijfluik)
9.  Plundering I — Diagnose
10. Plundering II — Mechaniek
11. Plundering III — Afloop
12. Plundering IV — Politieke landschap
13. Plundering 0 — Eerst plukken, dan oordelen
14. De actor is de regel
15. De Singaporese Gevolgenkaart
16. De Zwitserse Gevolgenkaart
"""
from pathlib import Path
import json
from openpyxl import load_workbook

REPO = Path("/tmp/gh-repo")
XLSX = REPO / "nl/_data/vizier.xlsx"
ROUTES_NL = REPO / "nl/_data/tabellen/2_routes_nl.json"

# (code-suffix, naam, ondertitel, slug)
ARTIKELEN = [
    # bestaande 1.1.1.1 = De voedingslijn — niet opnieuw toevoegen, alleen url-route bijwerken
    ("1.1.1.2", "Een staat met een auto-immuunziekte",
     "Een beeld om mee te denken, niet om mee te slaan — de metafoor die de Brussel-stukken draagt.",
     "de-metafoor-auto-immuunziekte"),

    ("1.1.1.3", "De anti-immuunziekte van Brussel maakt dat we arm worden",
     "CBAM, ETS en Pillar Two — drie aanvallen tegelijk. De BiCRS-route wordt genegeerd.",
     "de-anti-immuunziekte-van-brussel"),

    ("1.1.1.4", "Zij doden hun levensaders",
     "De diagnose in zes sectoren: één leiding, één anti-immuunziekte.",
     "zij-doden-hun-levensaders-brussel"),

    ("1.1.1.5", "De anti-immuunziekte van onze overheid slaat toe",
     "Den Haag, stikstof, 26 juni — de spiegel van Brussel in eigen tuin.",
     "de-auto-immuunziekte-slaat-toe"),

    ("1.1.1.6", "Zij doden hun levensader",
     "Den Haag-diagnose: boeren, MKB, productie. €212.795 per kilogram stikstof.",
     "zij-doden-hun-levensader"),

    ("1.1.1.7", "Het immuunsysteem dat gezonde cellen aanvalt",
     "Opinie · de medische metafoor toegepast op de Nederlandse staat.",
     "het-immuunsysteem-van-organisaties"),

    ("1.1.1.8", "De Grote Plundering",
     "Vijfluik over hoe Europa zijn productieven leegrooft, voordat de lichten uitgaan.",
     "de-grote-plundering"),

    ("1.1.1.9", "Plundering I — Diagnose",
     "Deel 1: wij rupfen wie ons voedt.",
     "plundering-1-diagnose"),

    ("1.1.1.10", "Plundering II — Mechaniek",
     "Deel 2: hoe het systeem zijn eigen producenten ontmoedigt.",
     "plundering-2-mechaniek"),

    ("1.1.1.11", "Plundering III — Afloop",
     "Deel 3: wat er resteert wanneer de productieven vertrokken zijn.",
     "plundering-3-afloop"),

    ("1.1.1.12", "Plundering IV — Politieke landschap",
     "Deel 4: hoe partij na partij meegaat in dezelfde fout.",
     "plundering-4-politieke-landschap"),

    ("1.1.1.13", "Plundering 0 — Eerst plukken, dan oordelen",
     "Proloog op de tetralogie: de logica van het oordeel-eerst.",
     "plundering-0-eerst-plukken-dan-oordelen"),

    ("1.1.1.14", "De actor is de regel",
     "Sluitstuk van de tetralogie — vier lagen van uitdrijving, één route eruit.",
     "de-actor-is-de-regel"),

    ("1.1.1.15", "De Singaporese Gevolgenkaart",
     "Wat Singapore doet wel goed en Brussel niet.",
     "de-singaporese-gevolgenkaart"),

    ("1.1.1.16", "De Zwitserse Gevolgenkaart",
     "Wat Zwitserland doet dat Brussel niet kan.",
     "de-zwitserse-gevolgenkaart"),
]

# =============================================================================
# Stap 1: XLSX bijwerken
# =============================================================================
wb = load_workbook(XLSX)
ws = wb['Knopen']

# Headers in rij 2
headers = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column+1)]
def col(name):
    return headers.index(name) + 1 if name in headers else None

c_code  = col("Code")
c_ouder = col("Ouder")
c_taal  = col("Taal")
c_type  = col("Type")
c_naam  = col("Naam")
c_ondt  = col("Ondertitel")
c_hero  = col("Hero")
c_herofilter = col("Hero-filter")
c_status = col("Status")
c_volg   = col("Volgorde")
c_actief = col("Actief")
c_auteur = col("Auteur")
c_url    = col("URL")

# Verzamel bestaande codes voor idempotentie
existing_codes = set()
last_row = 2
for r in range(3, ws.max_row + 1):
    code = ws.cell(r, c_code).value
    if code:
        existing_codes.add(str(code))
        last_row = r

# Update bestaande 1.1.1.1 volgorde naar 6 (was 1) zodat nieuwe items 1..16 logische volgorde krijgen
for r in range(3, ws.max_row + 1):
    if str(ws.cell(r, c_code).value) == "1.1.1.1":
        ws.cell(r, c_volg).value = 6
        # Zorg dat de URL leeg blijft, route komt uit JSON
        break

# Voeg nieuwe rijen toe
toegevoegd = 0
for i, (code, naam, ondt, slug) in enumerate(ARTIKELEN, start=1):
    if code in existing_codes:
        continue  # idempotent
    last_row += 1
    # Bepaal volgorde — proloog op 1, dan in volgorde 2..16 (6 = voedingslijn-rij die we hierboven al hebben)
    # Wij vullen 1, 2, 3, 4, 5, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16 voor de nieuwe items
    natuurlijke_volgorde = i if i < 6 else i + 1  # skip 6 (= voedingslijn)
    ws.cell(last_row, c_code).value = code
    ws.cell(last_row, c_ouder).value = "1.1.1"
    ws.cell(last_row, c_taal).value = "nl"
    ws.cell(last_row, c_type).value = "artikel"
    ws.cell(last_row, c_naam).value = naam
    ws.cell(last_row, c_ondt).value = ondt
    ws.cell(last_row, c_herofilter).value = "standaard"
    ws.cell(last_row, c_status).value = "live"
    ws.cell(last_row, c_volg).value = natuurlijke_volgorde
    ws.cell(last_row, c_actief).value = True
    ws.cell(last_row, c_auteur).value = "Jacobus van Merksteijn"
    toegevoegd += 1

wb.save(XLSX)
print(f"XLSX: {toegevoegd} rijen toegevoegd onder 1.1.1")

# =============================================================================
# Stap 2: 2_routes_nl.json bijwerken
# =============================================================================
routes = json.loads(ROUTES_NL.read_text(encoding="utf-8"))
existing_route_codes = {r["code"] for r in routes["rijen"]}

# Voor voedingslijn-bestand (1.1.1.1) - bestandsnaam is anders dan slug
URLS_TO_ADD = {
    "1.1.1.1": "wat-opkomt/de-methodologie-van-de-aanpassing.html",
}
for code, naam, ondt, slug in ARTIKELEN:
    URLS_TO_ADD[code] = f"wat-opkomt/{slug}.html"

# Voor elke nieuwe code: verifieer dat het bestand bestaat en voeg toe of update
toegevoegd_route = 0
geupdate_route = 0
existing_route_map = {r["code"]: r for r in routes["rijen"]}

for code, url in URLS_TO_ADD.items():
    fpath = REPO / "nl" / url
    if not fpath.exists():
        print(f"  WAARSCHUWING: bestand ontbreekt voor {code} → {url}")
        continue
    if code in existing_route_map:
        if existing_route_map[code].get("url") != url:
            existing_route_map[code]["url"] = url
            geupdate_route += 1
    else:
        # Maak nieuwe rij met alle JSON-velden (leeg waar onbekend)
        new_row = {key: None for key in routes["rijen"][0].keys()} if routes["rijen"] else {}
        new_row["code"] = code
        new_row["url"] = url
        new_row["doel_open"] = "zelf"
        routes["rijen"].append(new_row)
        toegevoegd_route += 1

ROUTES_NL.write_text(json.dumps(routes, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"2_routes_nl.json: {toegevoegd_route} routes toegevoegd, {geupdate_route} bijgewerkt")
print(f"Totaal routes nu: {len(routes['rijen'])}")
