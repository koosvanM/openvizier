#!/usr/bin/env python3
"""
Vertaal de navigatie-knopen voor ES/FR/IT/PT in vizier.xlsx.
Prefixen: FR=5, ES=6, IT=7, PT=8.
"""
from pathlib import Path
from openpyxl import load_workbook

REPO = Path("/tmp/gh-repo")
XLSX = REPO / "nl/_data/vizier.xlsx"

NAVIGATIE = {
    "1":      {"es": ("Hojear",              "qué hay nuevo"),
               "fr": ("Feuilleter",          "quoi de neuf"),
               "it": ("Sfogliare",           "cosa c'è di nuovo"),
               "pt": ("Folhear",             "o que há de novo")},
    "1.1":    {"es": ("Últimos artículos",   "cronológico"),
               "fr": ("Derniers articles",   "chronologique"),
               "it": ("Ultimi articoli",     "cronologico"),
               "pt": ("Últimos artigos",     "cronológico")},
    "1.2":    {"es": ("La portada",          "hoy"),
               "fr": ("La une",              "aujourd'hui"),
               "it": ("La prima pagina",     "oggi"),
               "pt": ("A capa",              "hoje")},
    "1.3":    {"es": ("Ediciones",           "ordenado por tema"),
               "fr": ("Éditions",            "classé par thème"),
               "it": ("Edizioni",            "ordinato per tema"),
               "pt": ("Edições",             "ordenado por tema")},
    "2":      {"es": ("Profundizar",         "los análisis"),
               "fr": ("Approfondir",         "les analyses"),
               "it": ("Approfondire",        "le analisi"),
               "pt": ("Aprofundar",          "as análises")},
    "2.1":    {"es": ("Mapas de consecuencias","el daño"),
               "fr": ("Cartes des conséquences","les dégâts"),
               "it": ("Mappe delle conseguenze","il danno"),
               "pt": ("Mapas de consequências","o dano")},
    "2.2":    {"es": ("BiCRS / Energía",     "la salida"),
               "fr": ("BiCRS / Énergie",     "la sortie"),
               "it": ("BiCRS / Energia",     "la via d'uscita"),
               "pt": ("BiCRS / Energia",     "a saída")},
    "2.3":    {"es": ("Política / Reforma",  "la forma"),
               "fr": ("Politique / Réforme", "la forme"),
               "it": ("Politica / Riforma",  "la forma"),
               "pt": ("Política / Reforma",  "a forma")},
    "2.4":    {"es": ("Países Bajos y mundo","el lugar"),
               "fr": ("Pays-Bas et monde",   "le lieu"),
               "it": ("Paesi Bassi e mondo", "il luogo"),
               "pt": ("Países Baixos e mundo","o lugar")},
    "3":      {"es": ("Investigar",          "dosieres y fuentes"),
               "fr": ("Rechercher",          "dossiers et sources"),
               "it": ("Ricercare",           "dossier e fonti"),
               "pt": ("Investigar",          "dossiês e fontes")},
    "3.1":    {"es": ("Dosieres",            "por país/tema"),
               "fr": ("Dossiers",            "par pays/thème"),
               "it": ("Dossier",             "per paese/tema"),
               "pt": ("Dossiês",             "por país/tema")},
    "3.2":    {"es": ("Investigación",       "fuentes primarias"),
               "fr": ("Recherche",           "sources primaires"),
               "it": ("Ricerca",             "fonti primarie"),
               "pt": ("Investigação",        "fontes primárias")},
    "3.3":    {"es": ("Archivo",             "releer todo"),
               "fr": ("Archives",            "tout relire"),
               "it": ("Archivio",            "rileggere tutto"),
               "pt": ("Arquivo",             "reler tudo")},
    "3.4":    {"es": ("Colofón",             "quién lo hace"),
               "fr": ("Colophon",            "qui le fait"),
               "it": ("Colophon",            "chi lo fa"),
               "pt": ("Colofão",             "quem o faz")},
}

PREFIX = {"fr": "5", "es": "6", "it": "7", "pt": "8"}

wb = load_workbook(XLSX)
ws = wb['Knopen']
headers = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column+1)]
def col(name): return headers.index(name) + 1 if name in headers else None
c_code, c_ouder, c_taal, c_type, c_naam, c_ondt = col("Code"), col("Ouder"), col("Taal"), col("Type"), col("Naam"), col("Ondertitel")

# Index huidige rijen
code_lang_to_row = {}
for r in range(3, ws.max_row + 1):
    code = ws.cell(r, c_code).value
    taal = ws.cell(r, c_taal).value
    if code: code_lang_to_row[(str(code), taal)] = r

bijgewerkt = toegevoegd = 0
last_row = ws.max_row

for suffix, trans in NAVIGATIE.items():
    for lang, (naam, ondt) in trans.items():
        full_code = f"{PREFIX[lang]}.{suffix}"
        key = (full_code, lang)
        if key in code_lang_to_row:
            r = code_lang_to_row[key]
            ws.cell(r, c_naam).value = naam
            ws.cell(r, c_ondt).value = ondt
            bijgewerkt += 1
        else:
            last_row += 1
            ws.cell(last_row, c_code).value = full_code
            ouder_parts = full_code.split(".")
            ws.cell(last_row, c_ouder).value = ".".join(ouder_parts[:-1]) if len(ouder_parts) > 1 else None
            ws.cell(last_row, c_taal).value = lang
            ws.cell(last_row, c_type).value = "ingang" if suffix in ("1","2","3") else "onderwerp"
            ws.cell(last_row, c_naam).value = naam
            ws.cell(last_row, c_ondt).value = ondt
            toegevoegd += 1

wb.save(XLSX)
print(f"Bijgewerkt: {bijgewerkt}, Toegevoegd: {toegevoegd}")
