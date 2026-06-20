#!/usr/bin/env python3
"""
Vertaal de navigatie-knopen (Bladeren / Verdiepen / Onderzoeken en hun
kinderen) in vizier.xlsx voor DE/EN/RU.

Codes (X = taal-prefix: 2=DE, 3=EN, 4=RU):
  X.1            Bladeren
  X.1.1          Laatste artikelen
  X.1.2          De voorpagina
  X.1.3          Edities
  X.2            Verdiepen
  X.2.1          Gevolgenkaarten
  X.2.2          BiCRS / Energie
  X.2.3          Politiek / Hervorming
  X.2.4          Nederland & Wereld
  X.3            Onderzoeken
  X.3.1          Dossiers
  X.3.2          Onderzoek
  X.3.3          Archief
  X.3.4          Colofon
"""
from pathlib import Path
from openpyxl import load_workbook

REPO = Path("/tmp/gh-repo")
XLSX = REPO / "nl/_data/vizier.xlsx"

# Vertalingen per code-suffix (zonder taal-prefix)
NAVIGATIE = {
    "1":      {"de": ("Blättern",          "was gibt es Neues"),
               "en": ("Browse",            "what is new"),
               "ru": ("Листать",           "что нового")},
    "1.1":    {"de": ("Neueste Artikel",   "chronologisch"),
               "en": ("Latest articles",   "chronological"),
               "ru": ("Последние статьи",  "хронологически")},
    "1.2":    {"de": ("Die Titelseite",    "heute"),
               "en": ("The front page",    "today"),
               "ru": ("Главная страница",  "сегодня")},
    "1.3":    {"de": ("Ausgaben",          "thematisch geordnet"),
               "en": ("Editions",          "thematically ordered"),
               "ru": ("Выпуски",           "тематически")},
    "2":      {"de": ("Vertiefen",         "die Analysen"),
               "en": ("Deepen",            "the analyses"),
               "ru": ("Углубиться",        "аналитика")},
    "2.1":    {"de": ("Folgenkarten",      "der Schaden"),
               "en": ("Consequence maps",  "the damage"),
               "ru": ("Карты последствий", "ущерб")},
    "2.2":    {"de": ("BiCRS / Energie",   "der Ausweg"),
               "en": ("BiCRS / Energy",    "the way out"),
               "ru": ("BiCRS / Энергия",   "путь наружу")},
    "2.3":    {"de": ("Politik / Reform",  "die Form"),
               "en": ("Politics / Reform", "the form"),
               "ru": ("Политика / Реформа","форма")},
    "2.4":    {"de": ("Niederlande & Welt","der Ort"),
               "en": ("Netherlands & World","the place"),
               "ru": ("Нидерланды и мир",  "место")},
    "3":      {"de": ("Recherchieren",     "Dossiers und Quellen"),
               "en": ("Research",          "dossiers and sources"),
               "ru": ("Исследовать",       "досье и источники")},
    "3.1":    {"de": ("Dossiers",          "nach Land/Thema"),
               "en": ("Dossiers",          "by country/theme"),
               "ru": ("Досье",             "по странам/темам")},
    "3.2":    {"de": ("Forschung",         "Primärquellen"),
               "en": ("Research",          "primary sources"),
               "ru": ("Исследования",      "первоисточники")},
    "3.3":    {"de": ("Archiv",            "alles nachlesen"),
               "en": ("Archive",           "read everything back"),
               "ru": ("Архив",             "перечитать всё")},
    "3.4":    {"de": ("Impressum",         "wer macht das"),
               "en": ("Colophon",          "who makes this"),
               "ru": ("Выходные данные",   "кто делает")},
}

# Taal-prefix
PREFIX = {"de": "2", "en": "3", "ru": "4"}

wb = load_workbook(XLSX)
ws = wb['Knopen']

# Headers
headers = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column+1)]
def col(name): return headers.index(name) + 1 if name in headers else None
c_code  = col("Code")
c_ouder = col("Ouder")
c_taal  = col("Taal")
c_type  = col("Type")
c_naam  = col("Naam")
c_ondt  = col("Ondertitel")

# Bouw look-up: (code, taal) -> rij
code_lang_to_row = {}
for r in range(3, ws.max_row + 1):
    code = ws.cell(r, c_code).value
    taal = ws.cell(r, c_taal).value
    if code:
        code_lang_to_row[(str(code), taal)] = r

bijgewerkt = 0
toegevoegd = 0
last_row = ws.max_row

for suffix, trans in NAVIGATIE.items():
    for lang, (naam, ondt) in trans.items():
        full_code = f"{PREFIX[lang]}.{suffix}"
        key = (full_code, lang)
        if key in code_lang_to_row:
            r = code_lang_to_row[key]
            old_naam = ws.cell(r, c_naam).value
            if old_naam != naam:
                ws.cell(r, c_naam).value = naam
                bijgewerkt += 1
            ws.cell(r, c_ondt).value = ondt
        else:
            # Nieuwe rij toevoegen
            last_row += 1
            ws.cell(last_row, c_code).value  = full_code
            # Ouder = code zonder laatste segment
            ouder_parts = full_code.split(".")
            ouder = ".".join(ouder_parts[:-1]) if len(ouder_parts) > 1 else None
            ws.cell(last_row, c_ouder).value = ouder
            ws.cell(last_row, c_taal).value  = lang
            ws.cell(last_row, c_type).value  = "ingang" if suffix in ("1","2","3") else "onderwerp"
            ws.cell(last_row, c_naam).value  = naam
            ws.cell(last_row, c_ondt).value  = ondt
            toegevoegd += 1

wb.save(XLSX)
print(f"Bijgewerkt: {bijgewerkt}, Toegevoegd: {toegevoegd}")
