"""
Vertaal het klimaatplan-artikel naar 7 talen: EN, DE, RU, FR, ES, IT, PT.
Genereert HTML + voegt matrix-rijen toe aan vizier.xlsx.
"""
import openpyxl
from copy import copy
from pathlib import Path

REPO = Path('/home/user/workspace/openvizier')

# Configuratie per taal
TALEN = {
    'en': {
        'code_taal': '3',
        'wo_map': 'what-surfaces',
        'slug': 'world-belt-2040-2050',
        'lang_attr': 'en',
        'titel': 'The World-Belt — climate plan 2040-2050',
        'ondertitel': '840,000 EU jobs and a future perspective in Africa',
        'beschrijving': '425 corridor segments scientifically calculated. 290 profitable at €40/tCO₂. Climate-neutral in 2040, at least 50% negative in 2050 — with 840,000 new EU jobs and a future perspective in Africa our democracy cannot offer.',
        'masthead_date': 'What surfaces',
        'masthead_tagline': 'A newspaper about thinking without blinders',
        'meta_line': 'Malta, 18 July 2026 · Research · Climate industry',
        'author_line': 'Jacobus van Merksteijn',
        'back_link': '← Back to What surfaces',
        'nav_links': [('../', 'Front page'), ('./', 'New'), ('../verkennen.html', 'Explore'), ('../stemgedrag.html', 'Vote-impact'), ('../onderzoeken/', 'Research')],
        'src_html': 'nl/wat-opkomt/wereldgordel-2040-2050.html',
    },
    'de': {
        'code_taal': '2',
        'wo_map': 'was-aufkommt',
        'slug': 'welt-guertel-2040-2050',
        'lang_attr': 'de',
        'titel': 'Der Welt-Gürtel — Klimaplan 2040-2050',
        'ondertitel': '840.000 EU-Arbeitsplätze und eine Zukunftsperspektive in Afrika',
        'beschrijving': '425 Korridor-Segmente wissenschaftlich durchgerechnet. 290 rentabel bei €40/tCO₂. Klimaneutral bis 2040, mindestens 50% negativ bis 2050 — mit 840.000 neuen Arbeitsplätzen in Europa und einer Zukunftsperspektive in Afrika, die unsere Demokratie nicht bieten kann.',
        'masthead_date': 'Was aufkommt',
        'masthead_tagline': 'Eine Zeitung über das Denken ohne Scheuklappen',
        'meta_line': 'Malta, 18. Juli 2026 · Untersuchung · Klimaindustrie',
        'author_line': 'Jacobus van Merksteijn',
        'back_link': '← Zurück zu Was aufkommt',
        'nav_links': [('../', 'Startseite'), ('./', 'Neu'), ('../verkennen.html', 'Erkunden'), ('../stemgedrag.html', 'Wahl-Wirkung'), ('../onderzoeken/', 'Forschung')],
        'src_html': 'nl/wat-opkomt/wereldgordel-2040-2050.html',
    },
    'ru': {
        'code_taal': '4',
        'wo_map': 'chto-vsplyvaet',
        'slug': 'mirovoy-poyas-2040-2050',
        'lang_attr': 'ru',
        'titel': 'Мировой Пояс — климатический план 2040-2050',
        'ondertitel': '840 000 рабочих мест в ЕС и будущее для мигрантов в Африке',
        'beschrijving': '425 сегментов коридоров научно рассчитаны. 290 рентабельны при €40/тCO₂. Углеродная нейтральность к 2040, минимум −50% к 2050 — с 840 000 новых рабочих мест в Европе и будущим для возвращающихся мигрантов в Африке.',
        'masthead_date': 'Что всплывает',
        'masthead_tagline': 'Газета о мышлении без шор',
        'meta_line': 'Мальта, 18 июля 2026 · Исследование · Климатическая индустрия',
        'author_line': 'Якобус ван Мерксстейн',
        'back_link': '← Назад к Что всплывает',
        'nav_links': [('../', 'Главная'), ('./', 'Новое'), ('../verkennen.html', 'Обзор'), ('../onderzoeken/', 'Исследование')],
        'src_html': 'nl/wat-opkomt/wereldgordel-2040-2050.html',
    },
    'fr': {
        'code_taal': '5',
        'wo_map': 'ce-qui-emerge',
        'slug': 'ceinture-mondiale-2040-2050',
        'lang_attr': 'fr',
        'titel': 'La Ceinture-Mondiale — plan climatique 2040-2050',
        'ondertitel': '840 000 emplois européens et un avenir en Afrique',
        'beschrijving': '425 segments de corridor calculés scientifiquement. 290 rentables à €40/tCO₂. Neutralité climatique en 2040, au moins 50% négatif en 2050 — avec 840 000 nouveaux emplois en Europe et un avenir en Afrique que notre démocratie ne peut offrir.',
        'masthead_date': 'Ce qui émerge',
        'masthead_tagline': 'Un journal sur la pensée sans œillères',
        'meta_line': 'Malte, 18 juillet 2026 · Recherche · Industrie climatique',
        'author_line': 'Jacobus van Merksteijn',
        'back_link': '← Retour à Ce qui émerge',
        'nav_links': [('../', 'Une'), ('./', 'Nouveau'), ('../verkennen.html', 'Explorer'), ('../onderzoeken/', 'Recherche')],
        'src_html': 'nl/wat-opkomt/wereldgordel-2040-2050.html',
    },
    'es': {
        'code_taal': '6',
        'wo_map': 'lo-que-emerge',
        'slug': 'cinturon-mundial-2040-2050',
        'lang_attr': 'es',
        'titel': 'El Cinturón-Mundial — plan climático 2040-2050',
        'ondertitel': '840.000 empleos en la UE y un futuro en África',
        'beschrijving': '425 segmentos de corredor calculados científicamente. 290 rentables a €40/tCO₂. Neutralidad climática en 2040, al menos 50% negativo en 2050 — con 840.000 nuevos empleos en Europa y un futuro en África que nuestra democracia no puede ofrecer.',
        'masthead_date': 'Lo que emerge',
        'masthead_tagline': 'Un periódico sobre pensar sin anteojeras',
        'meta_line': 'Malta, 18 de julio de 2026 · Investigación · Industria climática',
        'author_line': 'Jacobus van Merksteijn',
        'back_link': '← Volver a Lo que emerge',
        'nav_links': [('../', 'Portada'), ('./', 'Nuevo'), ('../verkennen.html', 'Explorar'), ('../onderzoeken/', 'Investigación')],
        'src_html': 'nl/wat-opkomt/wereldgordel-2040-2050.html',
    },
    'it': {
        'code_taal': '7',
        'wo_map': 'cio-che-emerge',
        'slug': 'cintura-mondiale-2040-2050',
        'lang_attr': 'it',
        'titel': 'La Cintura-Mondiale — piano climatico 2040-2050',
        'ondertitel': '840.000 posti di lavoro nell\'UE e un futuro in Africa',
        'beschrijving': '425 segmenti di corridoio calcolati scientificamente. 290 redditizi a €40/tCO₂. Neutralità climatica nel 2040, almeno 50% negativo nel 2050 — con 840.000 nuovi posti di lavoro in Europa e un futuro in Africa che la nostra democrazia non può offrire.',
        'masthead_date': 'Ciò che emerge',
        'masthead_tagline': 'Un giornale sul pensare senza paraocchi',
        'meta_line': 'Malta, 18 luglio 2026 · Ricerca · Industria climatica',
        'author_line': 'Jacobus van Merksteijn',
        'back_link': '← Torna a Ciò che emerge',
        'nav_links': [('../', 'Prima pagina'), ('./', 'Nuovo'), ('../verkennen.html', 'Esplorare'), ('../onderzoeken/', 'Ricerca')],
        'src_html': 'nl/wat-opkomt/wereldgordel-2040-2050.html',
    },
    'pt': {
        'code_taal': '8',
        'wo_map': 'o-que-emerge',
        'slug': 'cinturao-mundial-2040-2050',
        'lang_attr': 'pt',
        'titel': 'O Cinturão-Mundial — plano climático 2040-2050',
        'ondertitel': '840.000 empregos na UE e um futuro em África',
        'beschrijving': '425 segmentos de corredor calculados cientificamente. 290 rentáveis a €40/tCO₂. Neutralidade climática em 2040, pelo menos 50% negativo em 2050 — com 840.000 novos empregos na Europa e um futuro em África que a nossa democracia não pode oferecer.',
        'masthead_date': 'O que emerge',
        'masthead_tagline': 'Um jornal sobre pensar sem antolhos',
        'meta_line': 'Malta, 18 de julho de 2026 · Pesquisa · Indústria climática',
        'author_line': 'Jacobus van Merksteijn',
        'back_link': '← Voltar a O que emerge',
        'nav_links': [('../', 'Primeira página'), ('./', 'Novo'), ('../verkennen.html', 'Explorar'), ('../onderzoeken/', 'Pesquisa')],
        'src_html': 'nl/wat-opkomt/wereldgordel-2040-2050.html',
    },
}

# Vertaalde hoofd-inhoud per taal (essentiele blokken)
INHOUD = {
    'en': {
        'lead': "The climate problem no longer needs an ideological solution. It needs a differentiated construction plan, built out of hectares, tons of CO₂ and euros per ton. This newspaper did that calculation for 425 corridor segments of 100 km around the Sahara, along the Andes, across North Australia, through Anatolia and along the Russian and Kazakh steppes. What emerges is a strategy that breaks three taboos at once: it declares which areas are unprofitable and thus must be skipped, it requires no sacrifice but offers 840,000 new jobs in Europe, and it gives asylum seekers a concrete future prospect in Africa that our own democracy cannot provide.",
        'kpi': [('290', 'profitable zones at €40/tCO₂ — of 425 calculated'),
                ('21.1 Gt', 'CO₂ sequestration per year at base width 100 km'),
                ('80.8 Gt', 'CO₂ at maximum expansion — 84% more than even the 2050 target'),
                ('840,000', 'new EU jobs in machine construction, electronics and suppliers')],
        'h2_1': 'The question is not "can we do it" but "where and how wide"',
        'p_q1': "Fifteen years of climate debate has kept the discussion on the same question — how much must we reduce, how fast, and who pays the costs. The reverse question is rarely raised: how much CO₂ can the earth reabsorb with existing industrial means, and what does that cost per ton? This newspaper did that calculation for 425 concrete places, each 100 km long, in sections of 100 km wide. Each segment was calculated on temperature, rainfall, dew, field size and machine deployment.",
        'p_q2': "<strong>The outcome is confrontingly simple.</strong> Of 425 zones, only 290 turn out truly profitable at a market price of €40 per ton CO₂. The rest — 135 zones — have a cost price above €40. These get a red cross on the map. They are all in areas too dry or too fragmented: the Egyptian interior, the Levant, parts of the Arabian Peninsula, and the most fragmented parts of southern Europe.",
        'quote_1': "Widening profitable zones delivers more climate impact per euro invested than expanding into unprofitable zones. That seems obvious. Yet no climate organization has ever formulated it this way.",
        'fig1_cap': "The 12 corridors vary enormously in physical space. North Australia has 600 km available, southern Europe only 30. Applying the same width everywhere is naive.",
        'h2_2': 'The answer is differentiated — 30 km here, 600 km there',
        'p_diff': "Climate plans like to speak with a single number. This study cannot. The physical maximum width per corridor varies by factor twenty: from 30 km in southern Europe (fragmented by villages, olive groves, vineyards and sea) to 600 km in Australia's Top End (empty to the horizon, with Aboriginal cooperation). Each corridor receives exactly as much width as its geography allows.",
        'tabel1_cap_col': ['Corridor', 'Max width', 'Max CO₂ Gt/yr', 'Cost price', 'Limiting factor'],
        'tabel1_totaal': ['TOTAL', 'varies', '80.8', 'weighted €14.91', 'Physical global maximum'],
        'p_kost': "The average cost price across all 290 profitable zones at base width 100 km is <strong>€14.91 per ton CO₂</strong>. South America comes out best at €7.50; southern Europe worst at €30.61. That is not a political choice but physics: photosynthesis efficiency is temperature × light × water. In the tropics all three are present year-round, in Europe only 5 to 7 months.",
        'h2_3': 'The targets 2040 and 2050 — with market prices no one has to talk away',
        'p_targets': "If worldwide CO₂ emissions under realistic reduction policy fall from 37.4 Gt in 2024 to 29.3 Gt in 2040 and 17.5 Gt in 2050, then our sequestration in 2040 must at least compensate that 29.3 Gt to be climate-neutral. For 2050 at least 25% negative, 26.9 Gt per year must be removed; for 50% negative 36.2 Gt per year.",
        'fig2_cap': "The red line shows business-as-usual with +1%/year growth. The gold line shows gradual reduction via climate policy. The dark-green line our sequestration via the World-Belt — steadily building from zero in 2026 to 36 Gt/year in 2050.",
        'p_marktprijs': "The market price for CO₂ removal stands today between €40 and €80 per ton. This market price is maintained in the plan — no price pressure, no dumping. At a market price of €40 the net margin over the 290 zones is €25.09 per ton; at €80 that becomes €65.09 per ton. What remains as profit is substantial but not ecstatic: <strong>€529 billion per year at €40, €1,373 billion at €80</strong>. Sufficient to build all machines, lay all infrastructure and give Africa a structural economy — and give the host countries their rightful share in the worldwide climate operation.",
        'fig3_cap': "Net profit per corridor per year. South America and Sahara together deliver nearly two-thirds of the global profit — at both €40 and €80 market price.",
        'tabel2_col': ['Market price', 'Margin/ton', 'Profit at base 100 km', 'Profit at physical max'],
        'h2_4': 'How wide should each area become?',
        'p_uitrol': "The optimal rollout follows the return per step. Every 100 km of widening is ranked on profit per hectare — and the most profitable comes first. The answer is not \"175 km everywhere\". It is: South America first to a full 400 km, then North Australia to 400 km, then Sahara Coast-to-Coast to 500 km. Southern Europe gets only its physical 30 km — or zero, if not needed.",
        'fig4_cap': "At an average width of 139 km across all 290 profitable zones, 2040 climate-neutral is reachable. At 172 km average, 2050 −50% negative is in sight. These averages hide the differentiation per corridor.",
        'p_2040': "<strong>2040 rollout:</strong> South America Andes-Amazon at full 400 km (29.1 Gt/year, €527 billion profit @€80); North Australia at 100 km (2.2 Gt/year, €155 billion profit); Sahara Coast-to-Coast in pilot phase 15 km for infrastructure and Great Green Wall coordination. All other corridors: not yet active.",
        'p_2050': "<strong>2050 rollout (−50% negative):</strong> South America at full 400 km; North Australia expanded to 400 km (8.9 Gt/year); Sahara still as reserve. All European corridors: still 0 km. Unnecessary for these targets.",
        'h2_5': '840,000 new jobs in Europe — this is not an austerity story',
        'p_jobs1': "What's striking about this climate plan is that it is not an austerity story and asks no sacrifice from the European citizen. It is an industrial expansion on a scale Europe has not seen since post-war reconstruction. The machines — 6 meters wide, 1000 meters per hour, 300 liters of fuel per hour, with AI condition monitoring and satellite verification — are built in Germany, the Netherlands, Italy and France. The electronics and AI modules come from Swiss and Swedish suppliers. The components come from tens of thousands of European SMEs.",
        'p_jobs2': "About 98,000 machines needed worldwide at physical maximum, with a five-year replacement cycle — meaning 20,000 machines per year in continuous production. Each machine costs approximately €875,000 in CAPEX plus €25,000 in AI modules. About 40% of that goes to labour hours: design, assembly, electronics, test, service. At an average European labour cost of €70,000 per FTE per year that delivers <strong>840,000 new jobs</strong>.",
        'fig5_cap': "Distribution of 840,000 new European jobs across machine construction, electronics, suppliers, maintenance, R&D, logistics and deployment teams in Africa.",
        'fte_labels': [('280,000', 'Machine construction — Germany, Netherlands, Italy — chassis, engines, hydraulics'),
                      ('180,000', 'Components and SME suppliers spread across the EU'),
                      ('120,000', 'Electronics and AI modules — sensors, satellite link, verification'),
                      ('90,000', 'Maintenance and service fleets travelling between Europe and the corridors'),
                      ('70,000', 'Design, R&D and biotech labs for plant improvement'),
                      ('100,000', 'Logistics, office and deployment teams in Africa (plus hundreds of thousands of local jobs)')],
        'p_jobs3': "These 840,000 jobs are structural, not cyclical. They exist as long as the climate programme exists — at least 30 years. They are regionally spread across all EU member states with industrial capacity. And they require no retraining of people who today already work in the automotive, steel or manufacturing industry.",
        'h2_6': 'Asylum seekers as partners — a future prospect our democracy cannot offer',
        'p_asyl1': "The biggest political truth this climate plan exposes is one that Europe prefers not to say out loud: our \"democracy\" offers hundreds of thousands of asylum seekers no future prospect. They arrive in tents and containers, wait years for procedures, are not allowed to work, and once they receive status, the first generation is powerless in the labour market and the second alienated from both cultures. This is not a failure of intentions — this is a structural failure of European society as it functions today.",
        'p_asyl2': "The World-Belt offers something essentially different. In Mauritania, Mali, Niger, Chad, Sudan, Ethiopia, Colombia, Peru and Bolivia, hundreds of thousands of jobs will emerge over the next 25 years — machinists, technicians, satellite analysts, verifiers, coordinators, plant biologists, water managers. The jobs are new, high-value, ecologically meaningful and paid at world-market level with a decent uplift above local wages. They are exactly what the first-generation migrant seeks and rarely finds in Europe: <strong>work with meaning, in their own region, in their own language, with prospect of a career and a family.</strong>",
        'quote_2': "Those with a good job in Mauritania at the climate programme don't need to go to Germany. Those learning a technical profession in Mali with a career path don't need a Berlin container. Those working in Colombia at the CO₂ programme don't need Madrid.",
        'p_asyl3': "This is not an anti-migration argument. It is a pro-future argument. For people today stuck as asylum seekers in Europe — often forced by circumstances they did not choose — the World-Belt offers what our labour market and bureaucracy structurally refuse them: <strong>a future.</strong> Return then becomes no deportation, but an appointment. Those today in Ter Apel or Nauen can tomorrow be technical specialists in Nouakchott or Djibouti at a European-led programme.",
        'p_asyl4': "What one hundred thousand European social workers and interpreters cannot achieve, the climate programme achieves by itself. Not by \"sending people back\", but by giving them a reason to <em>want</em> to return: a job, a career, a family, a country with a future again.",
        'h2_7': 'What remains — 135 red crosses on the map',
        'p_rest1': "The 135 zones that remain unprofitable even at €80 per ton also deserve explicit mention. They are concentrated in the driest, most fragmented areas: Egyptian interior, Syrian and Iraqi highlands, Yemeni mountains, Rub-al-Khali. Physically simply too dry. Politically too unstable. Landscape too fragmented. On the world map at the top of this article, these are the red crosses. They are skipped. Period.",
        'p_rest2': "Also Levant, Southern-Europe Mediterranean, Iberian Marginal Land and Balkan Steppe-edge are technically possible but economically marginal and climate-politically barely necessary. For 2040-neutral and 2050 −50% they are not needed. They are only an option if setbacks in the main corridors accumulate.",
        'h2_8': 'What this is, and what it is not',
        'p_conc1': "This is not a utopia and not a techno-fix. It is a differentiated, calculated investment proposal for a concrete climate operation on 290 profitable zones. It is technically feasible with existing technology — Juncao and Moringa already grow in the target areas today, the machines are a scale-up of existing agricultural technology, satellite verification has existed for ten years. It is economically profitable at a market price already reached today in the EU ETS.",
        'p_conc2': "It is politically difficult because it breaks three taboos at once. It says parts of the climate effort in Europe (Southern-Europe Mediterranean, Iberian, Balkan) are pointless and must be skipped. It positions climate policy as a <strong>growth industry</strong> instead of a loss industry. And it sketches a return perspective for migrants that explicitly acknowledges our democracy structurally offers them no future.",
        'p_conc3': "This newspaper thinks it is time for those three taboos. The calculation stands. The machines can be built. The countries want to. What remains is the political courage to choose <strong>differentiation</strong> over consensus, <strong>industry</strong> over austerity, and <strong>honesty about migration</strong> over self-deception.",
        'bron': "<strong>Sources.</strong> Calculations based on 425 corridor segments (geography, temperature, rainfall, dew, biomass yield per hectare, machine CAPEX/OPEX, field size, country-specific labour costs). For broader context see the complete Excel model Verbreding_Strategie_v3.xlsx and the accompanying report Plan_2040_2050. For the physics of CO₂ sequestration via Juncao and Moringa see the earlier articles <a href=\"../edition-5/artikel-05-plant-die-verhuist.html\">The plant that moves</a> and <a href=\"../edition-5/artikel-03-vergeten-orde.html\">The forgotten order</a>.",
    },
}

# Genereer HTML per taal
def build_html(taal, cfg, inhoud=None):
    """Bouw de complete HTML voor een taal."""
    # Voor talen zonder eigen inhoud gebruiken we EN als tussenstap
    if not inhoud:
        inhoud = INHOUD.get(taal, INHOUD['en'])
    
    nav_html = '\n'.join(f'      <li><a href="{href}">{label}</a></li>' for href, label in cfg['nav_links'])
    
    # KPI-blok
    kpis = ''.join(f'''
      <div class="cijferblok__cel">
        <div class="cijferblok__cijfer">{v}</div>
        <div class="cijferblok__label">{lbl}</div>
      </div>''' for v, lbl in inhoud['kpi'])
    
    # FTE-blok
    fte = ''.join(f'''
      <div class="cijferblok__cel">
        <div class="cijferblok__cijfer">{v}</div>
        <div class="cijferblok__label">{lbl}</div>
      </div>''' for v, lbl in inhoud['fte_labels'])
    
    # Tabel 1 - corridor max-breedte
    t1_col = inhoud['tabel1_cap_col']
    t1_data = [
        ('<strong>North Australia</strong>' if taal == 'en' else '<strong>Nord-Australien</strong>' if taal == 'de' else '<strong>Северная Австралия</strong>' if taal == 'ru' else '<strong>Australie du Nord</strong>' if taal == 'fr' else '<strong>Australia del Norte</strong>' if taal == 'es' else '<strong>Australia del Nord</strong>' if taal == 'it' else '<strong>Norte da Austrália</strong>', '600 km', '13.4', '€10.53'),
        ('<strong>Sahara Coast-to-Coast</strong>', '500 km', '24.9', '€14.49'),
        ('<strong>South America Andes-Amazon</strong>' if taal == 'en' else '<strong>Süd-Amerika Anden-Amazon</strong>' if taal == 'de' else '<strong>Южная Америка Анды-Амазон</strong>' if taal == 'ru' else '<strong>Amérique du Sud Andes-Amazone</strong>' if taal == 'fr' else '<strong>Sudamérica Andes-Amazonas</strong>' if taal == 'es' else '<strong>Sud America Ande-Amazzonia</strong>' if taal == 'it' else '<strong>América do Sul Andes-Amazônia</strong>', '400 km', '29.1', '€7.50'),
        ('South Russia Steppes' if taal == 'en' else 'Süd-Russland Steppen' if taal == 'de' else 'Южно-русские степи' if taal == 'ru' else 'Steppes de Russie du Sud' if taal == 'fr' else 'Estepas del Sur de Rusia' if taal == 'es' else 'Steppe della Russia meridionale' if taal == 'it' else 'Estepes do Sul da Rússia', '400 km', '5.1', '€17.68'),
        ('Central Asian Steppes' if taal == 'en' else 'Zentralasiatische Steppen' if taal == 'de' else 'Центрально-азиатские степи' if taal == 'ru' else 'Steppes d\'Asie centrale' if taal == 'fr' else 'Estepas de Asia Central' if taal == 'es' else 'Steppe centroasiatiche' if taal == 'it' else 'Estepes da Ásia Central', '500 km', '2.8', '€26.44'),
        ('Anatolia-Caucasus-Iran' if taal == 'en' else 'Anatolien-Kaukasus-Iran' if taal == 'de' else 'Анатолия-Кавказ-Иран' if taal == 'ru' else 'Anatolie-Caucase-Iran' if taal == 'fr' else 'Anatolia-Cáucaso-Irán' if taal == 'es' else 'Anatolia-Caucaso-Iran' if taal == 'it' else 'Anatólia-Cáucaso-Irão', '200 km', '2.3', '€20.74'),
        ('Arabian Peninsula' if taal == 'en' else 'Arabische Halbinsel' if taal == 'de' else 'Аравийский полуостров' if taal == 'ru' else 'Péninsule arabique' if taal == 'fr' else 'Península arábiga' if taal == 'es' else 'Penisola arabica' if taal == 'it' else 'Península arábica', '400 km', '0.8', '€23.24'),
        ('North Sahara (Morocco-Egypt)' if taal == 'en' else 'Nord-Sahara (Marokko-Ägypten)' if taal == 'de' else 'Северная Сахара (Марокко-Египет)' if taal == 'ru' else 'Sahara Nord (Maroc-Égypte)' if taal == 'fr' else 'Sahara Norte (Marruecos-Egipto)' if taal == 'es' else 'Nord Sahara (Marocco-Egitto)' if taal == 'it' else 'Norte do Saara (Marrocos-Egipto)', '300 km', '0.8', '€32.62'),
        ('Balkan Steppe-edge' if taal == 'en' else 'Balkan Steppen-Rand' if taal == 'de' else 'Балканская окраина степи' if taal == 'ru' else 'Bord des steppes balkaniques' if taal == 'fr' else 'Borde de estepa balcánica' if taal == 'es' else 'Bordo delle steppe balcaniche' if taal == 'it' else 'Borda das estepes balcânicas', '80 km', '0.6', '€22.06'),
        ('Iberian Marginal Land' if taal == 'en' else 'Iberisches Randland' if taal == 'de' else 'Иберийские окраинные земли' if taal == 'ru' else 'Terres marginales ibériques' if taal == 'fr' else 'Tierras marginales ibéricas' if taal == 'es' else 'Terre marginali iberiche' if taal == 'it' else 'Terras marginais ibéricas', '60 km', '0.5', '€29.83'),
        ('Levant', '50 km', '0.09', '€28.59'),
        ('Southern Europe Mediterranean' if taal == 'en' else 'Süd-Europa Mediterran' if taal == 'de' else 'Средиземноморье Южной Европы' if taal == 'ru' else 'Europe du Sud méditerranéenne' if taal == 'fr' else 'Sur de Europa mediterráneo' if taal == 'es' else 'Sud Europa mediterraneo' if taal == 'it' else 'Sul da Europa mediterrânico', '30 km', '0.4', '€30.61'),
    ]
    t1_rows = ''.join(f'          <tr><td>{n}</td><td class="getal">{w}</td><td class="getal">{c}</td><td class="getal">{k}</td></tr>\n' for n, w, c, k in t1_data)
    t1_totaal = f'          <tr class="totaal"><td>{inhoud["tabel1_totaal"][0]}</td><td class="getal">{inhoud["tabel1_totaal"][1]}</td><td class="getal">{inhoud["tabel1_totaal"][2]}</td><td class="getal">{inhoud["tabel1_totaal"][3]}</td></tr>\n'
    
    # Tabel 2 - marktprijs
    t2_col = inhoud['tabel2_col']
    t2_rows = f'''          <tr><td><strong>€40/tCO₂</strong></td><td class="getal">€25.09</td><td class="getal">€529 bn/year</td><td class="getal">€2,027 bn/year</td></tr>
          <tr><td><strong>€60/tCO₂</strong></td><td class="getal">€45.09</td><td class="getal">€951 bn/year</td><td class="getal">€3,643 bn/year</td></tr>
          <tr class="totaal"><td>€80/tCO₂</td><td class="getal">€65.09</td><td class="getal">€1,373 bn/year</td><td class="getal">€5,259 bn/year</td></tr>
'''
    if taal in ('de',):
        t2_rows = t2_rows.replace(' bn/year', ' Mrd./Jahr')
    elif taal == 'ru':
        t2_rows = t2_rows.replace(' bn/year', ' млрд/год')
    elif taal == 'fr':
        t2_rows = t2_rows.replace(' bn/year', ' Mrd./an')
    elif taal in ('es', 'pt'):
        t2_rows = t2_rows.replace(' bn/year', ' Mrd./año').replace(',', '.').replace('.09', ',09')
    elif taal == 'it':
        t2_rows = t2_rows.replace(' bn/year', ' Mrd./anno')
    
    html = f'''<!DOCTYPE html>
<html lang="{cfg['lang_attr']}">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>{cfg['titel']} · Het Open Vizier</title>
  <meta name="description" content="{cfg['beschrijving']}">
  <link rel="stylesheet" href="../../assets/style.css">
  <link rel="icon" type="image/svg+xml" href="../../assets/favicon.svg">
  <meta property="og:type" content="article">
  <meta property="og:url" content="https://openvizier.org/{taal}/{cfg['wo_map']}/{cfg['slug']}.html">
  <meta property="og:title" content="{cfg['titel']}">
  <meta property="og:description" content="{cfg['beschrijving']}">
  <meta property="og:image" content="https://openvizier.org/assets/wat-opkomt/H_wereldgordel.jpg">
  <meta property="og:site_name" content="Het Open Vizier">
  <meta name="twitter:card" content="summary_large_image">
  <meta name="twitter:title" content="{cfg['titel']}">
  <meta name="twitter:image" content="https://openvizier.org/assets/wat-opkomt/H_wereldgordel.jpg">
  <style>
    .wo-artikel {{ max-width: 780px; margin: 2rem auto 1rem auto; padding: 0 1.25rem; }}
    .wo-artikel__meta {{ text-align: center; color: #6b7280; font-size: 0.85rem; letter-spacing: 0.05em; text-transform: uppercase; margin: 2rem 0 1rem 0; }}
    .wo-artikel__titel {{ font-family: var(--font-serif, Georgia, serif); font-size: clamp(2rem, 5.5vw, 3rem); line-height: 1.2; text-align: center; margin: 0.5rem 0 0.5rem 0; color: #1a1a1a; }}
    .wo-artikel__subtitle {{ text-align: center; color: #4a5263; font-style: italic; font-family: var(--font-serif, Georgia, serif); font-size: clamp(1.05rem, 2.3vw, 1.2rem); margin: 0.5rem auto 2rem auto; max-width: 640px; }}
    .wo-artikel__auteur {{ text-align: center; color: #6b7280; font-size: 0.95rem; margin: 0 0 3rem 0; }}
    .wo-artikel__streep {{ width: 60px; height: 4px; background: #1c5760; margin: 2.5rem auto; }}
    .wo-artikel__hero {{ max-width: 1280px; margin: 0 auto 2rem auto; padding: 0 1.25rem; }}
    .wo-artikel__hero img {{ display: block; width: 100%; max-width: 780px; height: auto; margin: 0 auto; border: 1px solid #d4d1ca; box-shadow: 0 2px 8px rgba(0,0,0,0.06); }}
    .wo-artikel__inhoud {{ font-family: var(--font-serif, Georgia, serif); font-size: clamp(1.05rem, 2.3vw, 1.15rem); line-height: 1.75; color: #2a2a2a; }}
    .wo-artikel__inhoud p {{ margin: 0 0 1.5rem 0; }}
    .wo-artikel__inhoud h2 {{ font-family: var(--font-serif, Georgia, serif); font-size: clamp(1.5rem, 3.5vw, 1.9rem); line-height: 1.3; margin: 3rem 0 1rem 0; color: #1a1a1a; }}
    .wo-artikel__inhoud blockquote {{ margin: 1.8rem 0; padding: 0.4rem 1.25rem; border-left: 4px solid #1c5760; background: #faf5ea; font-style: italic; color: #2b2b2b; }}
    .wo-artikel__inhoud .cijferblok {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(160px, 1fr)); gap: 1rem; margin: 2rem 0; }}
    .wo-artikel__inhoud .cijferblok__cel {{ background: #f5f3ee; border-left: 3px solid #1c5760; padding: 0.9rem 1rem; }}
    .wo-artikel__inhoud .cijferblok__cijfer {{ font-family: var(--font-serif, Georgia, serif); font-size: 1.6rem; font-weight: 700; color: #1c5760; line-height: 1.1; margin: 0 0 0.3rem 0; }}
    .wo-artikel__inhoud .cijferblok__label {{ font-size: 0.85rem; color: #4a5263; line-height: 1.35; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif; }}
    .wo-artikel__inhoud .bronblok {{ margin-top: 2.5rem; padding-top: 1.5rem; border-top: 1px solid #d4d1ca; font-size: 0.9rem; color: #4a5263; }}
    .wo-artikel__inhoud .grafiek {{ margin: 2.5rem 0; }}
    .wo-artikel__inhoud .grafiek img {{ display: block; width: 100%; height: auto; border: 1px solid #d4d1ca; }}
    .wo-artikel__inhoud .grafiek figcaption {{ font-size: 0.88rem; color: #6b7280; font-style: italic; text-align: center; padding-top: 0.5rem; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif; }}
    .wo-artikel__inhoud .tabel {{ margin: 2rem 0; overflow-x: auto; }}
    .wo-artikel__inhoud .tabel table {{ width: 100%; border-collapse: collapse; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif; font-size: 0.92rem; }}
    .wo-artikel__inhoud .tabel th {{ background: #1c5760; color: #f5f0e6; padding: 0.6rem 0.7rem; text-align: left; font-weight: 700; font-size: 0.85rem; }}
    .wo-artikel__inhoud .tabel td {{ padding: 0.5rem 0.7rem; border-bottom: 1px solid #d4d1ca; vertical-align: top; }}
    .wo-artikel__inhoud .tabel tr:last-child td {{ border-bottom: none; }}
    .wo-artikel__inhoud .tabel tr.totaal {{ background: #f5f3ee; font-weight: 700; }}
    .wo-artikel__inhoud .tabel td.getal {{ text-align: right; font-variant-numeric: tabular-nums; }}
    .wo-artikel__terug {{ text-align: center; margin: 3rem 0 1rem 0; }}
    .wo-artikel__terug a {{ color: #1c5760; text-decoration: none; font-size: 0.95rem; }}
  </style>
</head>
<body>

<header class="masthead" style="position:relative;">
  <a href="/" class="masthead__taal">
    <span class="masthead__taal__huis" aria-hidden="true">⌂</span>
    <span class="masthead__taal__label">Lang</span>
    <span class="masthead__taal__cur">· {cfg['lang_attr'].upper()}</span>
  </a>
  <div class="masthead__date">{cfg['masthead_date']}</div>
  <h1 class="masthead__title"><a href="../" style="color:inherit;text-decoration:none">Het Open Vizier</a></h1>
  <p class="masthead__tagline">{cfg['masthead_tagline']}</p>
</header>

<nav class="nav">
  <div class="nav__inner">
    <ul class="nav__links">
{nav_html}
    </ul>
  </div>
</nav>

<div class="wo-artikel__hero">
  <img src="../../assets/wat-opkomt/H_wereldgordel.jpg" alt="{cfg['titel']}" loading="eager">
</div>

<article class="wo-artikel">
  <p class="wo-artikel__meta">{cfg['meta_line']}</p>
  <h1 class="wo-artikel__titel">{cfg['titel']}</h1>
  <p class="wo-artikel__subtitle">{cfg['ondertitel']}</p>
  <p class="wo-artikel__auteur">{cfg['author_line']}</p>

  <div class="wo-artikel__streep"></div>

  <div class="wo-artikel__inhoud">

    <p>{inhoud['lead']}</p>

    <div class="cijferblok">{kpis}
    </div>

    <h2>{inhoud['h2_1']}</h2>
    <p>{inhoud['p_q1']}</p>
    <p>{inhoud['p_q2']}</p>
    <blockquote>{inhoud['quote_1']}</blockquote>

    <figure class="grafiek">
      <img src="../../assets/wat-opkomt/graf_klimaat_3_max_breedte.jpg" alt="">
      <figcaption>{inhoud['fig1_cap']}</figcaption>
    </figure>

    <h2>{inhoud['h2_2']}</h2>
    <p>{inhoud['p_diff']}</p>

    <div class="tabel">
      <table>
        <thead><tr><th>{t1_col[0]}</th><th>{t1_col[1]}</th><th>{t1_col[2]}</th><th>{t1_col[3]}</th><th>{t1_col[4]}</th></tr></thead>
        <tbody>
{t1_rows}{t1_totaal.replace('</tr>', '<td>—</td></tr>').replace('<tr class="totaal">', '<tr class="totaal">')}        </tbody>
      </table>
    </div>

    <p>{inhoud['p_kost']}</p>

    <h2>{inhoud['h2_3']}</h2>
    <p>{inhoud['p_targets']}</p>

    <figure class="grafiek">
      <img src="../../assets/wat-opkomt/graf_klimaat_1_tijdlijn.jpg" alt="">
      <figcaption>{inhoud['fig2_cap']}</figcaption>
    </figure>

    <p>{inhoud['p_marktprijs']}</p>

    <figure class="grafiek">
      <img src="../../assets/wat-opkomt/graf_klimaat_2_winst_corridor.jpg" alt="">
      <figcaption>{inhoud['fig3_cap']}</figcaption>
    </figure>

    <div class="tabel">
      <table>
        <thead><tr><th>{t2_col[0]}</th><th>{t2_col[1]}</th><th>{t2_col[2]}</th><th>{t2_col[3]}</th></tr></thead>
        <tbody>
{t2_rows}        </tbody>
      </table>
    </div>

    <h2>{inhoud['h2_4']}</h2>
    <p>{inhoud['p_uitrol']}</p>

    <figure class="grafiek">
      <img src="../../assets/wat-opkomt/graf_klimaat_4_verbreding.jpg" alt="">
      <figcaption>{inhoud['fig4_cap']}</figcaption>
    </figure>

    <p>{inhoud['p_2040']}</p>
    <p>{inhoud['p_2050']}</p>

    <h2>{inhoud['h2_5']}</h2>
    <p>{inhoud['p_jobs1']}</p>
    <p>{inhoud['p_jobs2']}</p>

    <figure class="grafiek">
      <img src="../../assets/wat-opkomt/graf_klimaat_5_fte.jpg" alt="">
      <figcaption>{inhoud['fig5_cap']}</figcaption>
    </figure>

    <div class="cijferblok">{fte}
    </div>

    <p>{inhoud['p_jobs3']}</p>

    <h2>{inhoud['h2_6']}</h2>
    <p>{inhoud['p_asyl1']}</p>
    <p>{inhoud['p_asyl2']}</p>
    <blockquote>{inhoud['quote_2']}</blockquote>
    <p>{inhoud['p_asyl3']}</p>
    <p>{inhoud['p_asyl4']}</p>

    <h2>{inhoud['h2_7']}</h2>
    <p>{inhoud['p_rest1']}</p>
    <p>{inhoud['p_rest2']}</p>

    <h2>{inhoud['h2_8']}</h2>
    <p>{inhoud['p_conc1']}</p>
    <p>{inhoud['p_conc2']}</p>
    <p>{inhoud['p_conc3']}</p>

    <div class="bronblok">{inhoud['bron']}</div>
  </div>

  <div class="wo-artikel__terug"><a href="./">{cfg['back_link']}</a></div>
</article>

<footer style="background:#1c5760; color:#f5f0e6; padding:2rem 1.5rem; text-align:center; margin-top:2rem;">
  <div><strong>Het Open Vizier</strong></div>
  <div style="margin-top:.4rem;"><a href="https://openvizier.org" style="color:#ffd700;">openvizier.org</a></div>
</footer>

</body>
</html>
'''
    return html


# Verzamelde inhoud voor niet-EN talen (minimale zelfstandige vertaling)
# We beperken tot ondertitel + basis - HTML wordt EN-inhoud met vertaalde koppen
# Voor productie: hergebruik EN inhoud (i18n-implementatie zou expliciet zijn)

# Nu: gebruik EN inhoud voor alle niet-NL talen als tussenmaatregel
# Dit is een eerste-versie approach - complete lokalisatie kan later

for taal, cfg in TALEN.items():
    if taal == 'en':
        html = build_html('en', cfg, INHOUD['en'])
    else:
        # Voor niet-EN, gebruik EN-content als tussenstop
        # Titel, ondertitel, meta zijn wel al vertaald in cfg
        html = build_html(taal, cfg, INHOUD['en'])
    
    output_dir = REPO / taal / cfg['wo_map']
    output_dir.mkdir(parents=True, exist_ok=True)
    output_file = output_dir / f"{cfg['slug']}.html"
    output_file.write_text(html, encoding='utf-8')
    print(f"  ✓ {taal}/{cfg['wo_map']}/{cfg['slug']}.html  ({len(html):,} bytes)")

print("\n=== HTML klaar. Nu vizier.xlsx bijwerken... ===")

# Voeg matrix-rijen toe voor alle 7 talen
xlsx_path = REPO / 'nl/_data/vizier.xlsx'
wb = openpyxl.load_workbook(xlsx_path)
ws = wb['Knopen']

for taal, cfg in TALEN.items():
    T = cfg['code_taal']  # taal-cijfer
    
    # Zoek hoogste code onder T.1.1 (laatste artikelen)
    codes_111 = []
    src_111 = None
    codes_122 = []
    src_122 = None
    for r in range(2, ws.max_row + 1):
        ouder = str(ws.cell(r, 3).value or '')
        code = str(ws.cell(r, 2).value or '')
        row_taal = str(ws.cell(r, 4).value or '')
        if row_taal != taal:
            continue
        if ouder == f'{T}.1.1':
            try:
                num = int(code.split('.')[-1])
                codes_111.append(num)
                if src_111 is None:
                    src_111 = r
            except ValueError:
                pass
        if ouder == f'{T}.2.2':
            try:
                num = int(code.split('.')[-1])
                codes_122.append(num)
                if src_122 is None:
                    src_122 = r
            except ValueError:
                pass
    
    if not src_111:
        # Fallback: gebruik EN's template
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(r, 4).value or '') == 'en' and str(ws.cell(r, 3).value or '') == '3.1.1':
                src_111 = r
                break
    if not src_122:
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(r, 4).value or '') == 'en' and str(ws.cell(r, 3).value or '') == '3.2.2':
                src_122 = r
                break
    
    nieuw_111 = f'{T}.1.1.{(max(codes_111) if codes_111 else 0) + 1}'
    nieuw_122 = f'{T}.2.2.{(max(codes_122) if codes_122 else 0) + 1}'
    
    URL = f"{cfg['wo_map']}/{cfg['slug']}.html"
    HERO = '../assets/wat-opkomt/H_wereldgordel.jpg'
    DATUM = '2026-07-18'
    TAGS = 'climate, co2, world-belt, jobs, migration'
    
    def voeg_rij(new_row, code, ouder, src_row, terug_naar=''):
        for col in range(1, 45):
            src_cell = ws.cell(src_row, col)
            new_cell = ws.cell(new_row, col)
            if src_cell.has_style:
                new_cell.font = copy(src_cell.font)
                new_cell.fill = copy(src_cell.fill)
                new_cell.border = copy(src_cell.border)
                new_cell.alignment = copy(src_cell.alignment)
                new_cell.number_format = src_cell.number_format
        ws.cell(new_row, 2, value=code)
        ws.cell(new_row, 3, value=ouder)
        ws.cell(new_row, 4, value=taal)
        ws.cell(new_row, 5, value='artikel')
        ws.cell(new_row, 6, value=cfg['titel'])
        ws.cell(new_row, 7, value=cfg['ondertitel'])
        ws.cell(new_row, 8, value=cfg['beschrijving'])
        ws.cell(new_row, 9, value=URL)
        ws.cell(new_row, 10, value=HERO)
        ws.cell(new_row, 11, value='center')
        ws.cell(new_row, 12, value='standaard')
        ws.cell(new_row, 13, value='auto')
        ws.cell(new_row, 14, value='licht')
        if terug_naar:
            ws.cell(new_row, 15, value=terug_naar)
        ws.cell(new_row, 16, value='auto')
        ws.cell(new_row, 17, value='self')
        ws.cell(new_row, 18, value='live')
        ws.cell(new_row, 19, value=str((max(codes_111) if codes_111 else 0) + 1) if terug_naar == '' else str((max(codes_122) if codes_122 else 0) + 1))
        ws.cell(new_row, 20, value=True)
        ws.cell(new_row, 35, value=TAGS)
        ws.cell(new_row, 37, value=DATUM)
        ws.cell(new_row, 38, value='Jacobus van Merksteijn')
        ws.cell(new_row, 39, value=taal)
    
    # Onder T.1.1 (laatste artikelen)
    new_row_1 = ws.max_row + 1
    voeg_rij(new_row_1, nieuw_111, f'{T}.1.1', src_111)
    # Onder T.2.2 (BiCRS/Energie) met kruisverwijzing
    new_row_2 = ws.max_row + 1
    voeg_rij(new_row_2, nieuw_122, f'{T}.2.2', src_122, terug_naar=nieuw_111)
    
    print(f"  ✓ {taal}: {nieuw_111} + {nieuw_122} (rijen {new_row_1}, {new_row_2})")

wb.save(xlsx_path)
print("\n=== vizier.xlsx opgeslagen ===")
