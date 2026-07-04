#!/usr/bin/env python3
"""
Harvest ES-vertalingen uit bestaande es_app/ door regel-voor-regel te vergelijken
met nl_app/ en de sleutel-lijst uit instructies.xlsx.

Voor elke rij in de matrix zoekt dit script naar de bijbehorende regel in
es_app/client/src/<bestand> en probeert de vertaalde tekst te extraheren
via dezelfde AST-methode. Als er een geldige ES-tekst is → vult ES-kolom.
Overige rijen blijven leeg en worden door een vertaal-subagent aangevuld.
"""
from __future__ import annotations
import html
import re
import subprocess
import json
from pathlib import Path
from openpyxl import load_workbook

WORKSPACE = Path("/home/user/workspace")
NL_APP = WORKSPACE / "nl_app"
ES_APP = WORKSPACE / "es_app"
MATRIX = WORKSPACE / "instructies.xlsx"


def normaliseer(tekst: str) -> str:
    """Normaliseer whitespace + entities + \\uXXXX-escapes."""
    if not tekst:
        return ""
    # \uXXXX → letterlijk
    tekst = re.sub(r"\\u([0-9a-fA-F]{4})", lambda m: chr(int(m.group(1), 16)), tekst)
    # HTML-entities → Unicode
    tekst = html.unescape(tekst)
    # Whitespace collapse
    return " ".join(tekst.split())


def bouw_regel_naar_es_map() -> dict[str, list[str]]:
    """Extraheer strings uit es_app/ per bestand via dezelfde AST-methode als de NL-extractor.
    Retourneert dict {bestand: [alle_es_strings_uit_dat_bestand]}.
    """
    # Draai de AST-extractor tegen es_app
    es_script = ES_APP / "extract_strings.mjs"
    # Kopieer NL-extractor script en pas SRC-pad aan
    bron = (NL_APP / "extract_strings.mjs").read_text()
    aangepast = bron.replace("/home/user/workspace/nl_app/", "/home/user/workspace/es_app/")
    # Filter uit: NL_MARKERS regex — bij ES-app willen we ALLE strings, ook Spaanse
    # Vervang de filter door "accepteer alles wat menselijk-tekst-achtig lijkt"
    aangepast = aangepast.replace(
        "if (NL_MARKERS.test(tekst)) return true;",
        "return true; // accepteer alles voor harvest"
    )
    es_script.write_text(aangepast)

    resultaat = subprocess.run(
        ["node", str(es_script)],
        capture_output=True, text=True, cwd=str(ES_APP),
    )
    if resultaat.returncode != 0:
        print(f"Extractie faalde: {resultaat.stderr}")
        return {}
    es_data = json.loads(resultaat.stdout)

    per_bestand: dict[str, list[dict]] = {}
    for item in es_data:
        per_bestand.setdefault(item["bestand"], []).append(item)
    return per_bestand


def harvest() -> None:
    wb = load_workbook(MATRIX)
    ws = wb["strings"]

    # Header zoek
    kopjes = [c.value for c in ws[1]]
    kol_nl = kopjes.index("nl") + 1
    kol_es = kopjes.index("es") + 1
    kol_bestand = kopjes.index("bestand") + 1
    kol_regel = kopjes.index("regel") + 1

    # ES-data extraheren
    print("→ ES-strings extraheren uit es_app/ (via AST)")
    es_per_bestand = bouw_regel_naar_es_map()
    if not es_per_bestand:
        print("❌ Geen ES-data geëxtraheerd.")
        return

    # Bouw regel→ES-string map per bestand
    regel_map: dict[str, dict[int, str]] = {}
    for bestand, items in es_per_bestand.items():
        regel_map[bestand] = {}
        for it in items:
            regel_map[bestand].setdefault(it["regel"], []).append(it["nl"])  # 'nl' key = geëxtraheerde tekst

    # Voor elke rij in matrix: zoek dichtstbijzijnde ES-string op zelfde regel
    gevuld = 0
    voor_vertaling = 0
    for rij in ws.iter_rows(min_row=2, max_row=ws.max_row):
        bestand = rij[kol_bestand - 1].value
        regel = rij[kol_regel - 1].value
        nl_tekst = rij[kol_nl - 1].value
        if not bestand or not nl_tekst:
            continue
        # Skip als ES al ingevuld
        if rij[kol_es - 1].value:
            continue

        # Zoek in es-map: eerst exacte regel, dan ±3 regels
        opties: list[str] = []
        for delta in [0, -1, 1, -2, 2, -3, 3]:
            r = (regel or 0) + delta
            als_lijst = regel_map.get(bestand, {}).get(r, [])
            opties.extend(als_lijst)

        # Filter: alleen niet-NL kandidaten (heuristiek: bevat geen typische NL-markers)
        nl_kort = normaliseer(nl_tekst)
        kandidaten = []
        for opt in opties:
            opt_norm = normaliseer(opt)
            if opt_norm == nl_kort:
                continue  # zelfde als NL = nog niet vertaald
            # Check: is dit waarschijnlijk Spaans? (bevat ¿ ¡ ñ, of eindigt op typische ES-woorden)
            if any(m in opt_norm.lower() for m in ["¿", "¡", "ción", "ente", "para", "los ", "las ", "una ", "está", "más", "país", "años"]):
                kandidaten.append(opt)
            # Ook: als lengte vergelijkbaar en niet identiek → accepteer
            elif 0.5 < len(opt_norm) / max(len(nl_kort), 1) < 2.0 and opt_norm != nl_kort:
                kandidaten.append(opt)

        if kandidaten:
            # Kies de langste kandidaat als beste match
            beste = max(kandidaten, key=len)
            rij[kol_es - 1].value = beste
            gevuld += 1
        else:
            voor_vertaling += 1

    wb.save(MATRIX)
    print(f"\n✅ {gevuld} rijen automatisch ingevuld met ES-tekst uit es_app/")
    print(f"⚠  {voor_vertaling} rijen wachten nog op vertaling")


if __name__ == "__main__":
    harvest()
