#!/usr/bin/env python3
"""
bouw_app.py <taal_code>

Genereert <taal>_app/ door de NL-app te kopiëren en elke geregistreerde
NL-string te vervangen door de vertaling uit instructies.xlsx.

Workflow:
  1. Lees instructies.xlsx (bron van waarheid voor vertalingen)
  2. Verifieer: voor elke rij, matcht kolom 'nl' met de tekst in bestand:regel
     - Zo niet: rapport + stop (bron uit sync)
  3. Kopieer nl_app/ naar <taal>_app/ (alleen source, geen node_modules/dist)
  4. Vervang in <taal>_app/ elke NL-string door <taal>-string
  5. Rapport: aantal vervangingen per bestand, ontbrekende vertalingen

Gebruik:
  python3 bouw_app.py es      # bouwt es_app/ uit nl_app/ + instructies.xlsx kolom 'es'
  python3 bouw_app.py de      # idem voor Duits
  python3 bouw_app.py --check # alleen verificatie, geen bouw
"""
from __future__ import annotations
import argparse
import html
import re
import shutil
import sys
from pathlib import Path
from openpyxl import load_workbook

WORKSPACE = Path("/home/user/workspace")
MATRIX = WORKSPACE / "instructies.xlsx"
JSON_KAART = WORKSPACE / "i18n_strings.json"  # bevat nl_exact voor patcher
NL_APP = WORKSPACE / "nl_app"

TALEN = {"de", "en", "es", "ch"}


def lees_matrix() -> list[dict]:
    """Lees instructies.xlsx → lijst van rij-dicts. Verrijk met nl_exact uit JSON-kaart."""
    import json as _json
    wb = load_workbook(MATRIX, data_only=True)
    ws = wb["strings"]
    kopjes = [c.value for c in ws[1]]
    rijen = []
    for rij in ws.iter_rows(min_row=2, values_only=True):
        if not rij[0]:
            continue
        rijen.append(dict(zip(kopjes, rij)))

    # Merge nl_exact uit JSON-kaart (sleutel als join-key)
    if JSON_KAART.exists():
        kaart = _json.loads(JSON_KAART.read_text())
        kaart_dict = {k["sleutel"]: k for k in kaart}
        for rij in rijen:
            k = kaart_dict.get(rij["sleutel"])
            if k:
                rij["nl_exact"] = k.get("nl_exact", rij["nl"])
            else:
                rij["nl_exact"] = rij["nl"]
    else:
        for rij in rijen:
            rij["nl_exact"] = rij["nl"]
    return rijen


def ontsnap_entities(tekst: str) -> str:
    """Vervang HTML/JSX entiteiten door hun Unicode-tegenhanger.
    Zowel &euml; als &#233; als losse escapes.
    """
    return html.unescape(tekst)


def verifieer_bron(rijen: list[dict]) -> tuple[int, list[str]]:
    """Check: elke matrix.nl moet exact voorkomen in bestand (na entity-normalisatie).
    Retourneert (ok_count, foutmeldingen).
    """
    ok = 0
    fouten: list[str] = []
    bestand_cache: dict[str, str] = {}  # genormaliseerde tekst
    for rij in rijen:
        bestand = rij["bestand"]
        nl_tekst = rij["nl"]
        if not nl_tekst:
            fouten.append(f"[LEGE NL] sleutel={rij['sleutel']} bestand={bestand}")
            continue
        if bestand not in bestand_cache:
            pad = NL_APP / "client" / "src" / bestand
            if not pad.exists():
                fouten.append(f"[BESTAND ONTBREEKT] {bestand}")
                continue
            ruw = pad.read_text(encoding="utf-8")
            # Unicode-escapes (\u00XX) → letterlijk Unicode karakter
            ruw = re.sub(r"\\u([0-9a-fA-F]{4})", lambda m: chr(int(m.group(1), 16)), ruw)
            # Entities → Unicode, dan whitespace normaliseren
            bestand_cache[bestand] = " ".join(ontsnap_entities(ruw).split())
        genormaliseerd_bestand = bestand_cache[bestand]
        genormaliseerd_zoek = " ".join(nl_tekst.split())
        if genormaliseerd_zoek in genormaliseerd_bestand:
            ok += 1
        else:
            fouten.append(f"[NIET GEVONDEN] {bestand}:{rij['regel']} sleutel={rij['sleutel']} tekst={nl_tekst[:60]!r}")
    return ok, fouten


def _zoek_varianten(nl_tekst: str, inhoud: str, nl_exact: str = None) -> list[str]:
    r"""Genereer zoekvarianten voor NL-tekst in tsx (entity-vormen + \uXXXX-escapes).
    nl_exact = letterlijke bron-tekst uit AST-extractie (behoudt entities/whitespace).
    Retourneert lijst in prioriteitsvolgorde.
    """
    varianten = []
    if nl_exact and nl_exact != nl_tekst:
        varianten.append(nl_exact)
    varianten.append(nl_tekst)
    # \uXXXX-escapes: elk niet-ASCII teken → \uXXXX (JavaScript-string-literal-vorm)
    escape_variant = "".join(
        c if ord(c) < 128 else f"\\u{ord(c):04x}"
        for c in nl_tekst
    )
    if escape_variant != nl_tekst:
        varianten.append(escape_variant)
    # Genereer entity-vorm: Unicode → HTML entity. LET OP: '&' NIET meemappen,
    # want dan corrumperen we alle andere entity-substituties.
    entity_map = _entity_map()
    entity_variant = nl_tekst
    for uni, ent in entity_map.items():
        entity_variant = entity_variant.replace(uni, ent)
    if entity_variant != nl_tekst:
        varianten.append(entity_variant)
    # Extra: als tekst apostrofs bevat, ook de &apos;-vorm proberen
    if "'" in entity_variant:
        varianten.append(entity_variant.replace("'", "&apos;"))
    if "'" in nl_tekst:
        varianten.append(nl_tekst.replace("'", "&apos;"))
    return varianten


def _entity_map() -> dict[str, str]:
    """Unicode-tekens → HTML entities, ZONDER '&' (die zou de conversie zelf breken)."""
    return {
        "ë": "&euml;", "ï": "&iuml;", "ö": "&ouml;", "ü": "&uuml;", "ä": "&auml;",
        "é": "&eacute;", "è": "&egrave;", "ê": "&ecirc;",
        "á": "&aacute;", "à": "&agrave;", "â": "&acirc;",
        "í": "&iacute;", "ì": "&igrave;", "î": "&icirc;",
        "ó": "&oacute;", "ò": "&ograve;", "ô": "&ocirc;",
        "ú": "&uacute;", "ù": "&ugrave;", "û": "&ucirc;",
        "ñ": "&ntilde;", "ç": "&ccedil;",
        "→": "&rarr;", "←": "&larr;", "↑": "&uarr;", "↓": "&darr;",
        "—": "&mdash;", "–": "&ndash;", "…": "&hellip;",
    }


def _veilige_vervanging(inhoud: str, zoek: str, vervang: str) -> str | None:
    """Vervang zoek door vervang, maar SLA OVER als match in code-context staat
    (import/export/from/function/class/const/let/var of geïdentificeerd als naam).
    Retourneert nieuwe inhoud of None als geen veilige match.
    """
    idx = 0
    while True:
        p = inhoud.find(zoek, idx)
        if p < 0:
            return None
        # Context: 40 tekens ervoor
        context_voor = inhoud[max(0, p - 40):p]
        # Als context_voor de laatste 'import ', 'export ', 'from ', 'function ', 'class ',
        # 'const ', 'let ', 'var ' bevat op de LAATSTE regel vóór p, is het code
        laatste_regel = context_voor.rsplit("\n", 1)[-1]
        code_context = any(
            kw in laatste_regel
            for kw in ["import ", "export ", "from ", "function ", "class ", " as "]
        )
        if code_context:
            idx = p + 1
            continue
        # Ook: als zoek direct voorafgegaan wordt door alfanumeriek of underscore, is het identifier-deel
        if p > 0 and (inhoud[p-1].isalnum() or inhoud[p-1] == "_"):
            idx = p + 1
            continue
        # En als zoek direct gevolgd wordt door alfanumeriek of underscore
        eind = p + len(zoek)
        if eind < len(inhoud) and (inhoud[eind].isalnum() or inhoud[eind] == "_"):
            idx = p + 1
            continue
        # Veilig — vervang deze en stop
        return inhoud[:p] + vervang + inhoud[eind:]


def _pas_entities_toe(bron_variant: str, vertaling: str) -> str:
    """Pas dezelfde entity-conventie toe op vertaling als op bron.
    Als bron entities gebruikte, converteer vertaling ook.
    """
    if "&" in bron_variant and ";" in bron_variant:
        # Bron gebruikt entities. Converteer speciale tekens in vertaling ook.
        for uni, ent in _entity_map().items():
            vertaling = vertaling.replace(uni, ent)
        vertaling = vertaling.replace("'", "&apos;") if "&apos;" in bron_variant else vertaling
    # Als bron \uXXXX-escapes gebruikt, converteer vertaling ook
    if re.search(r"\\u[0-9a-fA-F]{4}", bron_variant):
        vertaling = "".join(
            c if ord(c) < 128 else f"\\u{ord(c):04x}"
            for c in vertaling
        )
    return vertaling


def bouw(taal: str, rijen: list[dict]) -> None:
    """Kopieer nl_app → <taal>_app en vervang strings."""
    if taal not in TALEN:
        print(f"❌ Onbekende taal: {taal}. Gebruik één van: {sorted(TALEN)}")
        sys.exit(2)

    doel = WORKSPACE / f"{taal}_app"
    if doel.exists():
        print(f"→ Bestaande {doel.name} verwijderen")
        shutil.rmtree(doel)

    print(f"→ Kopiëren nl_app/ → {taal}_app/ (source only)")
    shutil.copytree(
        NL_APP, doel,
        ignore=shutil.ignore_patterns(
            "node_modules", "dist", ".vite", ".cache", "*.log",
            "extract_strings.mjs", "i18n_strings.json",
        ),
    )

    # Symlink node_modules van nl_app om npm install te vermijden
    (doel / "node_modules").symlink_to(NL_APP / "node_modules")

    # Per-bestand vervangingen accumuleren
    per_bestand: dict[str, list[tuple[str, str, str]]] = {}
    ontbrekend = 0
    for rij in rijen:
        vertaling = rij.get(taal)
        if not vertaling or str(vertaling).strip() == "":
            ontbrekend += 1
            continue
        per_bestand.setdefault(rij["bestand"], []).append(
            (rij["sleutel"], rij["nl"], rij.get("nl_exact", rij["nl"]), str(vertaling))
        )

    vervangingen_totaal = 0
    per_bestand_stats: dict[str, tuple[int, int]] = {}
    for bestand, items in per_bestand.items():
        pad = doel / "client" / "src" / bestand
        if not pad.exists():
            print(f"  ⚠ {bestand} niet gekopieerd?")
            continue
        inhoud = pad.read_text(encoding="utf-8")
        oorspronkelijk = inhoud
        # Normaliseer Unicode-escapes naar letterlijk Unicode voor betrouwbare zoek/vervang
        # (behoud oorspronkelijke voor niet-getroffen delen door alleen matchende delen te vervangen)
        # Sorteer op lengte aflopend om overlappende substrings juist te vervangen
        # Sorteer op lengte van nl_exact aflopend (langste eerst = overlap-veilig)
        items_gesorteerd = sorted(items, key=lambda x: -len(x[2]))
        raak = 0
        mis = 0
        for sleutel, nl_tekst, nl_exact, vert in items_gesorteerd:
            # Probeer varianten: exact (uit AST), genormaliseerd, entities, escapes
            gedaan = False
            for variant in _zoek_varianten(nl_tekst, inhoud, nl_exact):
                nieuwe_inhoud = _veilige_vervanging(inhoud, variant, _pas_entities_toe(variant, vert))
                if nieuwe_inhoud is not None:
                    inhoud = nieuwe_inhoud
                    raak += 1
                    gedaan = True
                    break
            if not gedaan:
                mis += 1
                print(f"  ⚠ {bestand}: kan {sleutel!r} niet vervangen (tekst niet gevonden)")
        if inhoud != oorspronkelijk:
            pad.write_text(inhoud, encoding="utf-8")
        per_bestand_stats[bestand] = (raak, mis)
        vervangingen_totaal += raak

    print()
    print(f"=== BOUW-RAPPORT {taal}_app ===")
    for b, (raak, mis) in sorted(per_bestand_stats.items()):
        vlag = "✅" if mis == 0 else "⚠"
        print(f"  {vlag} {b}: {raak} vervangen, {mis} gemist")
    print(f"\n  Totaal vervangen: {vervangingen_totaal}")
    print(f"  Ontbrekende vertalingen in matrix: {ontbrekend}")
    print(f"  Doelmap: {doel}")
    if ontbrekend > 0:
        print(f"\n⚠  {ontbrekend} sleutels hebben nog geen '{taal}'-vertaling in de matrix.")
        print(f"    Vul deze in {MATRIX.name} kolom '{taal}' en herstart bouw_app.py.")


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("taal", nargs="?", help=f"Taal-code: {sorted(TALEN)}")
    p.add_argument("--check", action="store_true", help="Alleen bron verifiëren, niet bouwen")
    args = p.parse_args()

    if not MATRIX.exists():
        print(f"❌ Matrix niet gevonden: {MATRIX}")
        sys.exit(1)
    if not NL_APP.exists():
        print(f"❌ NL-app niet gevonden: {NL_APP}")
        sys.exit(1)

    print(f"→ Lezen {MATRIX.name}")
    rijen = lees_matrix()
    print(f"  {len(rijen)} rijen")

    print(f"→ Verifiëren bron (matrix.nl ↔ nl_app/)")
    ok, fouten = verifieer_bron(rijen)
    print(f"  {ok}/{len(rijen)} rijen matchen exact")
    if fouten:
        print(f"\n❌ {len(fouten)} bron-mismatches:")
        for f in fouten[:20]:
            print(f"  {f}")
        if len(fouten) > 20:
            print(f"  ... en nog {len(fouten) - 20}")
        print("\n  → Fix instructies.xlsx of nl_app/ zodat matrix.nl overeenkomt met bestand.")
        sys.exit(1)

    if args.check:
        print("\n✅ Verificatie klaar. --check modus: geen bouw uitgevoerd.")
        return

    if not args.taal:
        p.error("taal is verplicht wanneer niet --check")

    bouw(args.taal, rijen)


if __name__ == "__main__":
    main()
